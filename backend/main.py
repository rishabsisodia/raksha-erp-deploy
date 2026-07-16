from fastapi import FastAPI, HTTPException, UploadFile, File, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response, HTMLResponse
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, ForeignKey, Text, text
from sqlalchemy.orm import declarative_base, sessionmaker, relationship
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, date
import os
import cloudinary
import cloudinary.uploader
import urllib.request
import csv
import io

app = FastAPI(title="Raksha ERP")

CLOUDINARY_URL = os.environ.get("CLOUDINARY_URL", "")
if CLOUDINARY_URL and "@" in CLOUDINARY_URL:
    parts = CLOUDINARY_URL.replace("cloudinary://", "").split("@")
    creds = parts[0].split(":")
    cloudinary.config(
        api_key=creds[0],
        api_secret=creds[1],
        cloud_name=parts[1],
        secure=True
    )

DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///./raksha_erp.db")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
if DATABASE_URL.startswith("postgresql://"):
    engine = create_engine(DATABASE_URL)
else:
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()


class Product(Base):
    __tablename__ = "products"
    id = Column(Integer, primary_key=True, index=True)
    part_no = Column(String, default="")
    name = Column(String, nullable=False)
    category = Column(String, default="")
    size = Column(String, default="")
    load_rating = Column(String, default="")
    material = Column(String, default="FRP")
    color = Column(String, default="Grey")
    unit = Column(String, default="Nos")
    hsn_code = Column(String, default="")
    pieces_per_box = Column(Integer, default=1)
    std_packaging = Column(Integer, default=1)
    created_at = Column(DateTime, default=datetime.utcnow)
    pricing = relationship("Pricing", back_populates="product", uselist=False, cascade="all,delete-orphan")


class Pricing(Base):
    __tablename__ = "pricing"
    id = Column(Integer, primary_key=True, index=True)
    product_id = Column(Integer, ForeignKey("products.id"), unique=True)
    raw_material_cost = Column(Float, default=0)
    labor_cost = Column(Float, default=0)
    overhead_cost = Column(Float, default=0)
    packing_cost = Column(Float, default=0)
    total_cost = Column(Float, default=0)
    profit_margin = Column(Float, default=20)
    mrp = Column(Float, default=0)
    dealer_price = Column(Float, default=0)
    distributor_price = Column(Float, default=0)
    gst_rate = Column(Float, default=18)
    product = relationship("Product", back_populates="pricing")


class Customer(Base):
    __tablename__ = "customers"
    id = Column(Integer, primary_key=True, index=True)
    customer_id = Column(String, unique=True, nullable=False)
    gstin = Column(String, default="")
    billing_address = Column(String, default="")
    shipping_address = Column(String, default="")
    state = Column(String, default="")
    district = Column(String, default="")
    city = Column(String, default="")
    pincode = Column(String, default="")
    contact_name = Column(String, default="")
    contact_number = Column(String, default="")
    contact_email = Column(String, default="")
    exec_code = Column(String, default="")
    exec_name = Column(String, default="")
    exec_number = Column(String, default="")
    exec_email = Column(String, default="")
    blacklisted = Column(Integer, default=0)
    created_at = Column(DateTime, default=datetime.utcnow)


class Transporter(Base):
    __tablename__ = "transporters"
    id = Column(Integer, primary_key=True, index=True)
    transporter_id = Column(String, unique=True, nullable=False)
    name = Column(String, nullable=False)
    phone = Column(String, default="")
    email = Column(String, default="")
    address = Column(Text, default="")
    state = Column(String, default="")
    district = Column(String, default="")
    city = Column(String, default="")
    pincode = Column(String, default="")
    gst_number = Column(String, default="")
    pan_number = Column(String, default="")
    gst_certificate = Column(String, default="")
    pan_card = Column(String, default="")
    contact_person = Column(String, default="")
    contact_number = Column(String, default="")
    blacklisted = Column(Integer, default=0)
    created_at = Column(DateTime, default=datetime.utcnow)


class Sale(Base):
    __tablename__ = "sales"
    id = Column(Integer, primary_key=True, index=True)
    invoice_no = Column(String)
    customer_id = Column(Integer, ForeignKey("customers.id"), nullable=True)
    product_id = Column(Integer, ForeignKey("products.id"), nullable=True)
    quantity = Column(Integer, nullable=True)
    unit_price = Column(Float, nullable=True)
    discount_percent = Column(Float, default=0)
    discount_amount = Column(Float, default=0)
    taxable_amount = Column(Float, nullable=True)
    cgst_rate = Column(Float, default=9)
    cgst_amount = Column(Float, default=0)
    sgst_rate = Column(Float, default=9)
    sgst_amount = Column(Float, default=0)
    freight_amount = Column(Float, default=0)
    total_amount = Column(Float, default=0)
    payment_status = Column(String, default="Pending")
    payment_method = Column(String, default="Cash")
    sale_date = Column(DateTime, nullable=True)
    notes = Column(String, default="")
    party_name = Column(String, default="")
    payment_terms = Column(String, default="")
    location = Column(String, default="")
    pincode = Column(String, default="")
    state = Column(String, default="")
    transporter_name = Column(String, default="")
    lr_no = Column(String, default="")
    weight_kgs = Column(Float, default=0)
    weight_pg_fiber = Column(Float, default=0)
    sales_person = Column(String, default="")
    pg_fiber_invoice_no = Column(String, default="")
    pg_fiber_invoice_value = Column(Float, default=0)
    gp = Column(Float, default=0)
    gp_percent = Column(Float, default=0)
    invoice_value = Column(Float, default=0)
    source_csv = Column(String, default="")
    customer = relationship("Customer")
    product = relationship("Product")


class Expense(Base):
    __tablename__ = "expenses"
    id = Column(Integer, primary_key=True, index=True)
    category = Column(String)
    description = Column(String, default="")
    amount = Column(Float)
    vendor = Column(String, default="")
    expense_date = Column(DateTime, default=datetime.utcnow)


class Settings(Base):
    __tablename__ = "settings"
    id = Column(Integer, primary_key=True, index=True)
    key = Column(String, unique=True)
    value = Column(String, default="")


class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, nullable=False)
    password_hash = Column(String, nullable=False)
    role = Column(String, default="user")
    created_at = Column(DateTime, default=datetime.utcnow)


class Token(Base):
    __tablename__ = "tokens"
    id = Column(Integer, primary_key=True, index=True)
    token = Column(String, unique=True, nullable=False)
    user_id = Column(Integer, ForeignKey("users.id"))
    created_at = Column(DateTime, default=datetime.utcnow)
    user = relationship("User")


class Order(Base):
    __tablename__ = "orders"
    id = Column(Integer, primary_key=True, index=True)
    sl_no = Column(Integer, default=0)
    po_no = Column(String, default="")
    po_date = Column(String, default="")
    customer_name = Column(String, default="")
    billing_site = Column(String, default="")
    shipping_site = Column(String, default="")
    no_of_boxes = Column(Integer, default=0)
    value_excl_gst_freight = Column(Float, default=0)
    invoice_no = Column(String, default="")
    invoice_date = Column(String, default="")
    invoice_amount_excl_gst = Column(Float, default=0)
    weight_kgs = Column(Float, default=0)
    freight_rate_per_kg = Column(Float, default=0)
    transport_charges = Column(Float, default=0)
    invoice_amount = Column(Float, default=0)
    eway_bill_no = Column(String, default="")
    lr_no = Column(String, default="")
    entry_date = Column(String, default="")
    credit_note_amount = Column(Float, default=0)
    credit_note_no = Column(String, default="")
    transporter = Column(String, default="")
    transporter_no = Column(String, default="")


class ProformaOrder(Base):
    __tablename__ = "proforma_orders"
    id = Column(Integer, primary_key=True, index=True)
    pi_no = Column(String, unique=True)
    pi_date = Column(DateTime, default=datetime.utcnow)
    customer_id = Column(Integer, ForeignKey("customers.id"))
    billing_site = Column(String, default="")
    shipping_site = Column(String, default="")
    no_of_boxes = Column(Integer, default=0)
    total_qty = Column(Integer, default=0)
    value_excl_gst = Column(Float, default=0)
    gst_amount = Column(Float, default=0)
    total_amount = Column(Float, default=0)
    freight_amount = Column(Float, default=0)
    payment_status = Column(String, default="Pending")
    payment_method = Column(String, default="Cash")
    transport_mode = Column(String, default="")
    delivery_days = Column(Integer, default=30)
    notes = Column(String, default="")
    terms = Column(Text, default="")
    order_type = Column(String, default="PI")
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    customer = relationship("Customer")
    items = relationship("ProformaOrderItem", back_populates="proforma_order", cascade="all,delete-orphan")


class ProformaOrderItem(Base):
    __tablename__ = "proforma_order_items"
    id = Column(Integer, primary_key=True, index=True)
    proforma_order_id = Column(Integer, ForeignKey("proforma_orders.id"))
    sl_no = Column(Integer)
    product_id = Column(Integer, ForeignKey("products.id"))
    part_no = Column(String, default="")
    description = Column(String, default="")
    size = Column(String, default="")
    category = Column(String, default="")
    qty_boxes = Column(Integer, default=1)
    std_packaging = Column(Integer, default=1)
    pieces_per_box = Column(Integer, default=1)
    final_qty = Column(Integer, default=0)
    mrp = Column(Float, default=0)
    d1 = Column(Float, default=0)
    d2 = Column(Float, default=0)
    d3 = Column(Float, default=0)
    d4 = Column(Float, default=0)
    d5 = Column(Float, default=0)
    cd = Column(Float, default=0)
    discount_percent = Column(Float, default=0)
    net_rate = Column(Float, default=0)
    lock_hinge = Column(Integer, default=0)
    basic_amount = Column(Float, default=0)
    created_at = Column(DateTime, default=datetime.utcnow)
    proforma_order = relationship("ProformaOrder", back_populates="items")
    product = relationship("Product")


@app.get("/api/db-info")
def db_info():
    db_url = os.environ.get("DATABASE_URL")
    has_key = "DATABASE_URL" in os.environ
    if db_url:
        masked = db_url[:20] + "...(masked)"
    else:
        masked = None
    return {
        "has_database_url_key": has_key,
        "db_url_preview": masked,
        "env_key_count": len(os.environ)
    }


@app.on_event("startup")
def startup_event():
    Base.metadata.create_all(bind=engine)
    with engine.connect() as conn:
        def safe_ddl(sql):
            try:
                conn.execute(text(sql))
                conn.commit()
            except Exception:
                conn.rollback()
        safe_ddl("CREATE TABLE IF NOT EXISTS proforma_orders (id SERIAL PRIMARY KEY, pi_no VARCHAR UNIQUE, pi_date TIMESTAMP, customer_id INTEGER REFERENCES customers(id), billing_site VARCHAR DEFAULT '', shipping_site VARCHAR DEFAULT '', no_of_boxes INTEGER DEFAULT 0, total_qty INTEGER DEFAULT 0, value_excl_gst FLOAT DEFAULT 0, gst_amount FLOAT DEFAULT 0, total_amount FLOAT DEFAULT 0, freight_amount FLOAT DEFAULT 0, payment_status VARCHAR DEFAULT 'Pending', payment_method VARCHAR DEFAULT 'Cash', transport_mode VARCHAR DEFAULT '', delivery_days INTEGER DEFAULT 30, notes VARCHAR DEFAULT '', terms TEXT DEFAULT '', order_type VARCHAR DEFAULT 'PI', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
        safe_ddl("CREATE TABLE IF NOT EXISTS proforma_order_items (id SERIAL PRIMARY KEY, proforma_order_id INTEGER REFERENCES proforma_orders(id), sl_no INTEGER, product_id INTEGER REFERENCES products(id), part_no VARCHAR DEFAULT '', description VARCHAR DEFAULT '', size VARCHAR DEFAULT '', category VARCHAR DEFAULT '', qty_boxes INTEGER DEFAULT 1, std_packaging INTEGER DEFAULT 1, pieces_per_box INTEGER DEFAULT 1, final_qty INTEGER DEFAULT 0, mrp FLOAT DEFAULT 0, d1 FLOAT DEFAULT 0, d2 FLOAT DEFAULT 0, d3 FLOAT DEFAULT 0, d4 FLOAT DEFAULT 0, d5 FLOAT DEFAULT 0, cd FLOAT DEFAULT 0, discount_percent FLOAT DEFAULT 0, net_rate FLOAT DEFAULT 0, lock_hinge INTEGER DEFAULT 0, basic_amount FLOAT DEFAULT 0, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
        safe_ddl("ALTER TABLE products ADD COLUMN IF NOT EXISTS part_no VARCHAR DEFAULT ''")
        safe_ddl("ALTER TABLE products ADD COLUMN IF NOT EXISTS pieces_per_box INTEGER DEFAULT 1")
        safe_ddl("ALTER TABLE products ADD COLUMN IF NOT EXISTS std_packaging INTEGER DEFAULT 1")
        safe_ddl("ALTER TABLE orders ADD COLUMN IF NOT EXISTS customer_name VARCHAR DEFAULT ''")
        safe_ddl("ALTER TABLE sales DROP CONSTRAINT IF EXISTS sales_invoice_no_key")
        new_sale_cols = [
            "party_name", "payment_terms", "location", "pincode", "state",
            "transporter_name", "lr_no", "weight_kgs", "weight_pg_fiber",
            "sales_person", "pg_fiber_invoice_no", "pg_fiber_invoice_value",
            "gp", "gp_percent", "source_csv"
        ]
        for col in new_sale_cols:
            col_type = "FLOAT" if col in ("weight_kgs","weight_pg_fiber","pg_fiber_invoice_value","gp","gp_percent") else "VARCHAR DEFAULT ''"
            safe_ddl(f"ALTER TABLE sales ADD COLUMN IF NOT EXISTS {col} {col_type}")
        safe_ddl("UPDATE sales SET invoice_value = '0' WHERE invoice_value IS NULL OR invoice_value = '' OR invoice_value = 'None' OR invoice_value = '\\u2013'")
        safe_ddl("ALTER TABLE sales ALTER COLUMN invoice_value TYPE FLOAT USING invoice_value::float")
        safe_ddl("ALTER TABLE sales ADD COLUMN IF NOT EXISTS invoice_value FLOAT DEFAULT 0")
        customer_cols = [
            ("customer_id", "VARCHAR DEFAULT ''"),
            ("gstin", "VARCHAR DEFAULT ''"),
            ("billing_address", "VARCHAR DEFAULT ''"),
            ("shipping_address", "VARCHAR DEFAULT ''"),
            ("state", "VARCHAR DEFAULT ''"),
            ("district", "VARCHAR DEFAULT ''"),
            ("city", "VARCHAR DEFAULT ''"),
            ("pincode", "VARCHAR DEFAULT ''"),
            ("contact_name", "VARCHAR DEFAULT ''"),
            ("contact_number", "VARCHAR DEFAULT ''"),
            ("contact_email", "VARCHAR DEFAULT ''"),
            ("exec_code", "VARCHAR DEFAULT ''"),
            ("exec_name", "VARCHAR DEFAULT ''"),
            ("exec_number", "VARCHAR DEFAULT ''"),
            ("exec_email", "VARCHAR DEFAULT ''"),
            ("blacklisted", "INTEGER DEFAULT 0"),
            ("created_at", "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
        ]
        for col_name, col_type in customer_cols:
            safe_ddl(f"ALTER TABLE customers ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
        safe_ddl("UPDATE customers SET customer_id = 'C' || id WHERE customer_id IS NULL OR customer_id = ''")
        safe_ddl("ALTER TABLE customers ADD CONSTRAINT customers_customer_id_unique UNIQUE (customer_id)")
    backfill_part_numbers()
    backfill_pieces_per_box()
    backfill_product_names()
    seed_data()


PIECES_PER_BOX_MAP = {
    "10x10": 12, "12x12": 6, "15x15": 6, "18x18": 4,
    "21x21": 3, "24x24": 2, "26x26": 1, "28x28": 1,
    "30x30": 1, "36x36": 1, "42x42": 1,
    "12x18": 6, "12x24": 5, "18x24": 3,
    "250x250": 12, "300x300": 6, "380x380": 6, "450x450": 4,
    "530x530": 3, "600x600": 2, "660x660": 1, "700x700": 1, "710x710": 1,
    "750x750": 1, "900x900": 1, "1060x1060": 1, "1065x1065": 1,
    "300x450": 6, "300x600": 5, "450x600": 3,
}


def backfill_pieces_per_box():
    db = SessionLocal()
    try:
        products = db.query(Product).all()
        updated = 0
        for p in products:
            size_lower = (p.size or "").lower().replace(" ", "")
            ppb = PIECES_PER_BOX_MAP.get(size_lower)
            if ppb and p.pieces_per_box != ppb:
                p.pieces_per_box = ppb
                p.std_packaging = ppb
                updated += 1
            pn = (p.part_no or "").upper().replace(" ", "")
            if pn.startswith("FRP012"):
                if p.load_rating != "10 Ton":
                    p.load_rating = "10 Ton"
                    updated += 1
            elif pn.startswith("FRP01") or pn.startswith("FRP04") or pn.startswith("FRP10"):
                if p.load_rating != "5 Ton" or not p.load_rating:
                    p.load_rating = "5 Ton"
                    updated += 1
        if updated:
            db.commit()
            print(f"Backfilled pieces_per_box and tonnage for {updated} products")
    finally:
        db.close()


PART_NO_CSV = [
    ("FRP01101-GRY", "Raksha FRP Manhole Cover 10x10 - 5 Ton Grey"),
    ("FRP01103-GRY", "Raksha FRP Manhole Cover 12x12 - 5 Ton Grey"),
    ("FRP01106-GRY", "Raksha FRP Manhole Cover 15x15 - 5 Ton Grey"),
    ("FRP01109-GRY", "Raksha FRP Manhole Cover 18x18 - 5 Ton Grey"),
    ("FRP01112-GRY", "Raksha FRP Manhole Cover 21x21 - 5 Ton Grey"),
    ("FRP01115-GRY", "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey"),
    ("FRP01117-GRY", "Raksha FRP Manhole Cover 26x26 - 5 Ton Grey"),
    ("FRP01119-GRY", "Raksha FRP Manhole Cover 28x28 - 5 Ton Grey"),
    ("FRP01121-GRY", "Raksha FRP Manhole Cover 30x30 - 5 Ton Grey"),
    ("FRP01127-GRY", "Raksha FRP Manhole Cover 36x36 - 5 Ton Grey"),
    ("FRP04106-GRY", "Raksha FRP Manhole Cover 12x18 - 5 Ton Grey"),
    ("FRP04112-GRY", "Raksha FRP Manhole Cover 12x24 - 5 Ton Grey"),
    ("FRP10106-GRY", "Raksha FRP Manhole Cover 18x24 - 5 Ton Grey"),
    ("FRP01101-WH", "Raksha FRP Manhole Cover 10x10 - 5 Ton White"),
    ("FRP01103-WH", "Raksha FRP Manhole Cover 12x12 - 5 Ton White"),
    ("FRP01106-WH", "Raksha FRP Manhole Cover 15x15 - 5 Ton White"),
    ("FRP01109-WH", "Raksha FRP Manhole Cover 18x18 - 5 Ton White"),
    ("FRP01112-WH", "Raksha FRP Manhole Cover 21x21 - 5 Ton White"),
    ("FRP01115-WH", "Raksha FRP Manhole Cover 24x24 - 5 Ton White"),
    ("FRP01117-WH", "Raksha FRP Manhole Cover 26x26 - 5 Ton White"),
    ("FRP01119-WH", "Raksha FRP Manhole Cover 28x28 - 5 Ton White"),
    ("FRP01121-WH", "Raksha FRP Manhole Cover 30x30 - 5 Ton White"),
    ("FRP01127-WH", "Raksha FRP Manhole Cover 36x36 - 5 Ton White"),
    ("FRP04106-WH", "Raksha FRP Manhole Cover 12x18 - 5 Ton White"),
    ("FRP04112-WH", "Raksha FRP Manhole Cover 12x24 - 5 Ton White"),
    ("FRP10106-WH", "Raksha FRP Manhole Cover 18x24 - 5 Ton White"),
    ("FRP01112-GRYL", "Raksha FRP Manhole Cover 21x21 - 5 Ton Grey (with Lock)"),
    ("FRP01115-GRYL", "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey (with Lock)"),
    ("FRP01117-GRYL", "Raksha FRP Manhole Cover 26x26 - 5 Ton Grey (with Lock)"),
    ("FRP01119-GRYL", "Raksha FRP Manhole Cover 28x28 - 5 Ton Grey (with Lock)"),
    ("FRP01121-GRYL", "Raksha FRP Manhole Cover 30x30 - 5 Ton Grey (with Lock)"),
    ("FRP01127-GRYL", "Raksha FRP Manhole Cover 36x36 - 5 Ton Grey (with Lock)"),
    ("FRP01112-WHL", "Raksha FRP Manhole Cover 21x21 - 5 Ton White (with Lock)"),
    ("FRP01115-WHL", "Raksha FRP Manhole Cover 24x24 - 5 Ton White (with Lock)"),
    ("FRP01117-WHL", "Raksha FRP Manhole Cover 26x26 - 5 Ton White (with Lock)"),
    ("FRP01119-WHL", "Raksha FRP Manhole Cover 28x28 - 5 Ton White (with Lock)"),
    ("FRP01121-WHL", "Raksha FRP Manhole Cover 30x30 - 5 Ton White (with Lock)"),
    ("FRP01127-WHL", "Raksha FRP Manhole Cover 36x36 - 5 Ton White (with Lock)"),
    ("FRP01115-GRYH", "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey (Double Hinges)"),
    ("FRP01115-WHH", "Raksha FRP Manhole Cover 24x24 - 5 Ton White (Double Hinges)"),
    ("FRP01115-GRY/H&L", "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey (Double Hinges & Lock)"),
    ("FRP01115-WH/H&L", "Raksha FRP Manhole Cover 24x24 - 5 Ton White (Double Hinges & Lock)"),
    ("FRP01209-GRY", "Raksha FRP Manhole Cover 18x18 - 10 Ton Grey"),
    ("FRP01215-GRY", "Raksha FRP Manhole Cover 24x24 - 10 Ton Grey"),
    ("FRP01219-GRY", "Raksha FRP Manhole Cover 28x28 - 10 Ton Grey"),
    ("FRP01221-GRY", "Raksha FRP Manhole Cover 30x30 - 10 Ton Grey"),
    ("FRP01233-GRY", "Raksha FRP Manhole Cover 42x42 - 10 Ton Grey"),
    ("FRP01209-WH", "Raksha FRP Manhole Cover 18x18 - 10 Ton White"),
    ("FRP01215-WH", "Raksha FRP Manhole Cover 24x24 - 10 Ton White"),
    ("FRP01219-WH", "Raksha FRP Manhole Cover 28x28 - 10 Ton White"),
    ("FRP01221-WH", "Raksha FRP Manhole Cover 30x30 - 10 Ton White"),
    ("FRP01233-WH", "Raksha FRP Manhole Cover 42x42 - 10 Ton White"),
    ("RGC00001-GRY", "Raksha FRP Gully Cover 10x10 - Grey"),
    ("RGC00002-GRY", "Raksha FRP Gully Cover 12x12 - Grey"),
    ("RGC00003-GRY", "Raksha FRP Gully Cover 15x15 - Grey"),
    ("RGC00004-GRY", "Raksha FRP Gully Cover 18x18 - Grey"),
    ("RGC00005-GRY", "Raksha FRP Gully Cover 24x24 - Grey"),
    ("RGC00001-WH", "Raksha FRP Gully Cover 10x10 - White"),
    ("RGC00002-WH", "Raksha FRP Gully Cover 12x12 - White"),
    ("RGC00003-WH", "Raksha FRP Gully Cover 15x15 - White"),
    ("RGC00004-WH", "Raksha FRP Gully Cover 18x18 - White"),
    ("RGC00005-WH", "Raksha FRP Gully Cover 24x24 - White"),
]


def generate_part_no(product):
    if product.part_no:
        return product.part_no
    import re
    name_lower = (product.name or "").lower()
    size = (product.size or "").lower().replace(" ", "")
    color = (product.color or "").lower()
    has_lock = "lock" in name_lower
    has_hinges = "hinge" in name_lower
    is_gully = "gully" in name_lower
    is_10ton = "10 ton" in name_lower

    size_dim_map = {
        "250x250": ("10x10", "10x10"), "300x300": ("12x12", "12x12"),
        "380x380": ("15x15", "15x15"), "450x450": ("18x18", "18x18"),
        "535x535": ("21x21", "21x21"), "600x600": ("24x24", "24x24"),
        "660x660": ("26x26", "26x26"), "710x710": ("28x28", "28x28"),
        "760x760": ("30x30", "30x30"), "900x900": ("36x36", "36x36"),
        "1065x1065": ("42x42", "42x42"),
        "300x450": ("12x18", "12x18"), "300x600": ("12x24", "12x24"),
        "450x600": ("18x24", "18x24"),
    }
    nominal = size_dim_map.get(size)
    if not nominal:
        m = re.search(r'(\d+)\s*x\s*(\d+)', product.size or product.name or "")
        if m:
            nominal = (f"{m.group(1)}x{m.group(2)}", f"{m.group(1)}x{m.group(2)}")
        else:
            return ""

    nom_lower = nominal[0]
    color_name = "White" if "white" in color else "Grey"
    tonnage = "10 Ton" if is_10ton else "5 Ton"

    suffix = ""
    if has_hinges and has_lock:
        suffix = " (Double Hinges & Lock)"
    elif has_hinges:
        suffix = " (Double Hinges)"
    elif has_lock:
        suffix = " (with Lock)"

    prefix = "RGC" if is_gully else "FRP"
    if is_gully:
        csv_desc = f"Raksha FRP Gully Cover {nom_lower} - {color_name}{suffix}"
    else:
        csv_desc = f"Raksha FRP Manhole Cover {nom_lower} - {tonnage} {color_name}{suffix}"

    for pn, desc in PART_NO_CSV:
        if desc.strip().lower() == csv_desc.strip().lower():
            return pn

    return ""


def backfill_part_numbers():
    db = SessionLocal()
    try:
        updated = 0
        for p in db.query(Product).filter((Product.part_no == "") | (Product.part_no.is_(None))).all():
            pn = generate_part_no(p)
            if pn:
                p.part_no = pn
                updated += 1
        if updated:
            db.commit()
            print(f"Backfilled part_no for {updated} products")
    finally:
        db.close()


def get_new_product_name(part_no, old_name=""):
    pn = (part_no or "").upper().replace(" ", "")
    is_gully = pn.startswith("RGC")
    is_10ton = pn.startswith("FRP012")

    suffix = ""
    if pn.endswith("/H&L"):
        suffix = " (Double Hinges & Lock)"
    elif pn.endswith("H") and not pn.endswith("-WH"):
        suffix = " (Double Hinges)"
    elif pn.endswith("L") and "GRY" not in pn[-4:]:
        suffix = " (with Lock)"

    color = "White" if "-WH" in pn else "Grey"

    size_map = {
        "FRP01101": "10x10", "FRP01103": "12x12", "FRP01106": "15x15",
        "FRP01109": "18x18", "FRP01112": "21x21", "FRP01115": "24x24",
        "FRP01117": "26x26", "FRP01119": "28x28", "FRP01121": "30x30",
        "FRP01127": "36x36", "FRP04106": "12x18", "FRP04112": "12x24",
        "FRP10106": "18x24", "FRP01209": "18x18", "FRP01215": "24x24",
        "FRP01219": "28x28", "FRP01221": "30x30", "FRP01233": "42x42",
        "RGC00001": "10x10", "RGC00002": "12x12", "RGC00003": "15x15",
        "RGC00004": "18x18", "RGC00005": "24x24",
    }

    size = ""
    for prefix, sz in size_map.items():
        if pn.startswith(prefix):
            size = sz
            break

    if not size:
        m = __import__("re").search(r"(\d+)\s*[xX]\s*(\d+)", old_name or "")
        if m:
            size = f"{m.group(1)}x{m.group(2)}"

    if is_gully:
        return f"Raksha FRP Gully Cover {size} - {color}{suffix}"

    tonnage = "10 Ton" if is_10ton else "5 Ton"
    return f"Raksha FRP Manhole Cover {size} - {tonnage} {color}{suffix}"


def backfill_product_names():
    db = SessionLocal()
    try:
        updated = 0
        for p in db.query(Product).all():
            new_name = get_new_product_name(p.part_no, p.name)
            if new_name and new_name != p.name:
                p.name = new_name
                updated += 1
        if updated:
            db.commit()
            print(f"Backfilled product names for {updated} products")
    finally:
        db.close()


def seed_data():
    db = SessionLocal()
    try:
        import hashlib
        admin = db.query(User).filter(User.username == "admin").first()
        if not admin:
            admin = User(username="admin", password_hash=hashlib.sha256("admin123".encode()).hexdigest(), role="admin")
            db.add(admin)
            db.commit()

        if db.query(Product).count() > 0:
            return

        products = [
            # Manhole Cover - Grey (5 Ton)
            {"part_no": "FRP01101-GRY", "name": "Raksha FRP Manhole Cover 10x10 - 5 Ton Grey", "category": "Manhole Cover", "size": "10x10", "color": "Grey", "rate": 190, "mrp": 686, "ppb": 12, "tonnage": "5 Ton"},
            {"part_no": "FRP01103-GRY", "name": "Raksha FRP Manhole Cover 12x12 - 5 Ton Grey", "category": "Manhole Cover", "size": "12x12", "color": "Grey", "rate": 242, "mrp": 830, "ppb": 6, "tonnage": "5 Ton"},
            {"part_no": "FRP01106-GRY", "name": "Raksha FRP Manhole Cover 15x15 - 5 Ton Grey", "category": "Manhole Cover", "size": "15x15", "color": "Grey", "rate": 310, "mrp": 1046, "ppb": 6, "tonnage": "5 Ton"},
            {"part_no": "FRP01109-GRY", "name": "Raksha FRP Manhole Cover 18x18 - 5 Ton Grey", "category": "Manhole Cover", "size": "18x18", "color": "Grey", "rate": 455, "mrp": 1536, "ppb": 4, "tonnage": "5 Ton"},
            {"part_no": "FRP01112-GRY", "name": "Raksha FRP Manhole Cover 21x21 - 5 Ton Grey", "category": "Manhole Cover", "size": "21x21", "color": "Grey", "rate": 640, "mrp": 2130, "ppb": 3, "tonnage": "5 Ton"},
            {"part_no": "FRP01115-GRY", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey", "category": "Manhole Cover", "size": "24x24", "color": "Grey", "rate": 765, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            {"part_no": "FRP01117-GRY", "name": "Raksha FRP Manhole Cover 26x26 - 5 Ton Grey", "category": "Manhole Cover", "size": "26x26", "color": "Grey", "rate": 1130, "mrp": 3266, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01119-GRY", "name": "Raksha FRP Manhole Cover 28x28 - 5 Ton Grey", "category": "Manhole Cover", "size": "28x28", "color": "Grey", "rate": 1500, "mrp": 4934, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01121-GRY", "name": "Raksha FRP Manhole Cover 30x30 - 5 Ton Grey", "category": "Manhole Cover", "size": "30x30", "color": "Grey", "rate": 1750, "mrp": 5854, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01127-GRY", "name": "Raksha FRP Manhole Cover 36x36 - 5 Ton Grey", "category": "Manhole Cover", "size": "36x36", "color": "Grey", "rate": 3200, "mrp": 11454, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP04106-GRY", "name": "Raksha FRP Manhole Cover 12x18 - 5 Ton Grey", "category": "Manhole Cover", "size": "12x18", "color": "Grey", "rate": 350, "mrp": 1154, "ppb": 6, "tonnage": "5 Ton"},
            {"part_no": "FRP04112-GRY", "name": "Raksha FRP Manhole Cover 12x24 - 5 Ton Grey", "category": "Manhole Cover", "size": "12x24", "color": "Grey", "rate": 500, "mrp": 1624, "ppb": 5, "tonnage": "5 Ton"},
            {"part_no": "FRP10106-GRY", "name": "Raksha FRP Manhole Cover 18x24 - 5 Ton Grey", "category": "Manhole Cover", "size": "18x24", "color": "Grey", "rate": 620, "mrp": 2020, "ppb": 3, "tonnage": "5 Ton"},
            # Manhole Cover - White (5 Ton)
            {"part_no": "FRP01101-WH", "name": "Raksha FRP Manhole Cover 10x10 - 5 Ton White", "category": "Manhole Cover", "size": "10x10", "color": "White", "rate": 190, "mrp": 686, "ppb": 12, "tonnage": "5 Ton"},
            {"part_no": "FRP01103-WH", "name": "Raksha FRP Manhole Cover 12x12 - 5 Ton White", "category": "Manhole Cover", "size": "12x12", "color": "White", "rate": 242, "mrp": 830, "ppb": 6, "tonnage": "5 Ton"},
            {"part_no": "FRP01106-WH", "name": "Raksha FRP Manhole Cover 15x15 - 5 Ton White", "category": "Manhole Cover", "size": "15x15", "color": "White", "rate": 310, "mrp": 1046, "ppb": 6, "tonnage": "5 Ton"},
            {"part_no": "FRP01109-WH", "name": "Raksha FRP Manhole Cover 18x18 - 5 Ton White", "category": "Manhole Cover", "size": "18x18", "color": "White", "rate": 455, "mrp": 1536, "ppb": 4, "tonnage": "5 Ton"},
            {"part_no": "FRP01112-WH", "name": "Raksha FRP Manhole Cover 21x21 - 5 Ton White", "category": "Manhole Cover", "size": "21x21", "color": "White", "rate": 640, "mrp": 2130, "ppb": 3, "tonnage": "5 Ton"},
            {"part_no": "FRP01115-WH", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton White", "category": "Manhole Cover", "size": "24x24", "color": "White", "rate": 765, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            {"part_no": "FRP01117-WH", "name": "Raksha FRP Manhole Cover 26x26 - 5 Ton White", "category": "Manhole Cover", "size": "26x26", "color": "White", "rate": 1130, "mrp": 3266, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01119-WH", "name": "Raksha FRP Manhole Cover 28x28 - 5 Ton White", "category": "Manhole Cover", "size": "28x28", "color": "White", "rate": 1500, "mrp": 4934, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01121-WH", "name": "Raksha FRP Manhole Cover 30x30 - 5 Ton White", "category": "Manhole Cover", "size": "30x30", "color": "White", "rate": 1750, "mrp": 5854, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01127-WH", "name": "Raksha FRP Manhole Cover 36x36 - 5 Ton White", "category": "Manhole Cover", "size": "36x36", "color": "White", "rate": 3200, "mrp": 11454, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP04106-WH", "name": "Raksha FRP Manhole Cover 12x18 - 5 Ton White", "category": "Manhole Cover", "size": "12x18", "color": "White", "rate": 350, "mrp": 1154, "ppb": 6, "tonnage": "5 Ton"},
            {"part_no": "FRP04112-WH", "name": "Raksha FRP Manhole Cover 12x24 - 5 Ton White", "category": "Manhole Cover", "size": "12x24", "color": "White", "rate": 500, "mrp": 1624, "ppb": 5, "tonnage": "5 Ton"},
            {"part_no": "FRP10106-WH", "name": "Raksha FRP Manhole Cover 18x24 - 5 Ton White", "category": "Manhole Cover", "size": "18x24", "color": "White", "rate": 620, "mrp": 2020, "ppb": 3, "tonnage": "5 Ton"},
            # Manhole Cover - Grey With Lock (5 Ton)
            {"part_no": "FRP01112-GRYL", "name": "Raksha FRP Manhole Cover 21x21 - 5 Ton Grey (with Lock)", "category": "Manhole Cover", "size": "21x21", "color": "Grey", "rate": 710, "mrp": 2130, "ppb": 3, "tonnage": "5 Ton"},
            {"part_no": "FRP01115-GRYL", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey (with Lock)", "category": "Manhole Cover", "size": "24x24", "color": "Grey", "rate": 835, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            {"part_no": "FRP01117-GRYL", "name": "Raksha FRP Manhole Cover 26x26 - 5 Ton Grey (with Lock)", "category": "Manhole Cover", "size": "26x26", "color": "Grey", "rate": 1270, "mrp": 3266, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01119-GRYL", "name": "Raksha FRP Manhole Cover 28x28 - 5 Ton Grey (with Lock)", "category": "Manhole Cover", "size": "28x28", "color": "Grey", "rate": 1640, "mrp": 4934, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01121-GRYL", "name": "Raksha FRP Manhole Cover 30x30 - 5 Ton Grey (with Lock)", "category": "Manhole Cover", "size": "30x30", "color": "Grey", "rate": 1890, "mrp": 5854, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01127-GRYL", "name": "Raksha FRP Manhole Cover 36x36 - 5 Ton Grey (with Lock)", "category": "Manhole Cover", "size": "36x36", "color": "Grey", "rate": 3340, "mrp": 11454, "ppb": 1, "tonnage": "5 Ton"},
            # Manhole Cover - White With Lock (5 Ton)
            {"part_no": "FRP01112-WHL", "name": "Raksha FRP Manhole Cover 21x21 - 5 Ton White (with Lock)", "category": "Manhole Cover", "size": "21x21", "color": "White", "rate": 710, "mrp": 2130, "ppb": 3, "tonnage": "5 Ton"},
            {"part_no": "FRP01115-WHL", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton White (with Lock)", "category": "Manhole Cover", "size": "24x24", "color": "White", "rate": 835, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            {"part_no": "FRP01117-WHL", "name": "Raksha FRP Manhole Cover 26x26 - 5 Ton White (with Lock)", "category": "Manhole Cover", "size": "26x26", "color": "White", "rate": 1270, "mrp": 3266, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01119-WHL", "name": "Raksha FRP Manhole Cover 28x28 - 5 Ton White (with Lock)", "category": "Manhole Cover", "size": "28x28", "color": "White", "rate": 1640, "mrp": 4934, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01121-WHL", "name": "Raksha FRP Manhole Cover 30x30 - 5 Ton White (with Lock)", "category": "Manhole Cover", "size": "30x30", "color": "White", "rate": 1890, "mrp": 5854, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01127-WHL", "name": "Raksha FRP Manhole Cover 36x36 - 5 Ton White (with Lock)", "category": "Manhole Cover", "size": "36x36", "color": "White", "rate": 3340, "mrp": 11454, "ppb": 1, "tonnage": "5 Ton"},
            # Manhole Cover - Double Hinges (5 Ton)
            {"part_no": "FRP01115-GRYH", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey (Double Hinges)", "category": "Manhole Cover", "size": "24x24", "color": "Grey", "rate": 965, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            {"part_no": "FRP01115-WHH", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton White (Double Hinges)", "category": "Manhole Cover", "size": "24x24", "color": "White", "rate": 965, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            # Manhole Cover - Double Hinges & Lock (5 Ton)
            {"part_no": "FRP01115-GRY/H&L", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey (Double Hinges & Lock)", "category": "Manhole Cover", "size": "24x24", "color": "Grey", "rate": 1065, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            {"part_no": "FRP01115-WH/H&L", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton White (Double Hinges & Lock)", "category": "Manhole Cover", "size": "24x24", "color": "White", "rate": 1065, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            # Heavy Duty Manhole Cover - Grey (10 Ton)
            {"part_no": "FRP01209-GRY", "name": "Raksha FRP Manhole Cover 18x18 - 10 Ton Grey", "category": "Manhole Cover", "size": "18x18", "color": "Grey", "rate": 1200, "mrp": 3340, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01215-GRY", "name": "Raksha FRP Manhole Cover 24x24 - 10 Ton Grey", "category": "Manhole Cover", "size": "24x24", "color": "Grey", "rate": 2200, "mrp": 6042, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01219-GRY", "name": "Raksha FRP Manhole Cover 28x28 - 10 Ton Grey", "category": "Manhole Cover", "size": "28x28", "color": "Grey", "rate": 3100, "mrp": 8504, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01221-GRY", "name": "Raksha FRP Manhole Cover 30x30 - 10 Ton Grey", "category": "Manhole Cover", "size": "30x30", "color": "Grey", "rate": 3800, "mrp": 10414, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01233-GRY", "name": "Raksha FRP Manhole Cover 42x42 - 10 Ton Grey", "category": "Manhole Cover", "size": "42x42", "color": "Grey", "rate": 11000, "mrp": 30208, "ppb": 1, "tonnage": "10 Ton"},
            # Heavy Duty Manhole Cover - White (10 Ton)
            {"part_no": "FRP01209-WH", "name": "Raksha FRP Manhole Cover 18x18 - 10 Ton White", "category": "Manhole Cover", "size": "18x18", "color": "White", "rate": 1200, "mrp": 3340, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01215-WH", "name": "Raksha FRP Manhole Cover 24x24 - 10 Ton White", "category": "Manhole Cover", "size": "24x24", "color": "White", "rate": 2200, "mrp": 6042, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01219-WH", "name": "Raksha FRP Manhole Cover 28x28 - 10 Ton White", "category": "Manhole Cover", "size": "28x28", "color": "White", "rate": 3100, "mrp": 8504, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01221-WH", "name": "Raksha FRP Manhole Cover 30x30 - 10 Ton White", "category": "Manhole Cover", "size": "30x30", "color": "White", "rate": 3800, "mrp": 10414, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01233-WH", "name": "Raksha FRP Manhole Cover 42x42 - 10 Ton White", "category": "Manhole Cover", "size": "42x42", "color": "White", "rate": 11000, "mrp": 30208, "ppb": 1, "tonnage": "10 Ton"},
            # Gully Cover - Grey
            {"part_no": "RGC00001-GRY", "name": "Raksha FRP Gully Cover 10x10 - Grey", "category": "Gully Cover", "size": "10x10", "color": "Grey", "rate": 240, "mrp": 806, "ppb": 12},
            {"part_no": "RGC00002-GRY", "name": "Raksha FRP Gully Cover 12x12 - Grey", "category": "Gully Cover", "size": "12x12", "color": "Grey", "rate": 325, "mrp": 984, "ppb": 6},
            {"part_no": "RGC00003-GRY", "name": "Raksha FRP Gully Cover 15x15 - Grey", "category": "Gully Cover", "size": "15x15", "color": "Grey", "rate": 440, "mrp": 1380, "ppb": 6},
            {"part_no": "RGC00004-GRY", "name": "Raksha FRP Gully Cover 18x18 - Grey", "category": "Gully Cover", "size": "18x18", "color": "Grey", "rate": 570, "mrp": 2012, "ppb": 4},
            {"part_no": "RGC00005-GRY", "name": "Raksha FRP Gully Cover 24x24 - Grey", "category": "Gully Cover", "size": "24x24", "color": "Grey", "rate": 1160, "mrp": 3910, "ppb": 2},
            # Gully Cover - White
            {"part_no": "RGC00001-WH", "name": "Raksha FRP Gully Cover 10x10 - White", "category": "Gully Cover", "size": "10x10", "color": "White", "rate": 240, "mrp": 806, "ppb": 12},
            {"part_no": "RGC00002-WH", "name": "Raksha FRP Gully Cover 12x12 - White", "category": "Gully Cover", "size": "12x12", "color": "White", "rate": 325, "mrp": 984, "ppb": 6},
            {"part_no": "RGC00003-WH", "name": "Raksha FRP Gully Cover 15x15 - White", "category": "Gully Cover", "size": "15x15", "color": "White", "rate": 440, "mrp": 1380, "ppb": 6},
            {"part_no": "RGC00004-WH", "name": "Raksha FRP Gully Cover 18x18 - White", "category": "Gully Cover", "size": "18x18", "color": "White", "rate": 570, "mrp": 2012, "ppb": 4},
            {"part_no": "RGC00005-WH", "name": "Raksha FRP Gully Cover 24x24 - White", "category": "Gully Cover", "size": "24x24", "color": "White", "rate": 1160, "mrp": 3910, "ppb": 2},
        ]

        pid = 1
        for prod in products:
            p = Product(id=pid, part_no=prod["part_no"], name=prod["name"], category=prod["category"], size=prod["size"], load_rating=prod.get("tonnage", "5 Ton"), material="FRP", color=prod["color"], hsn_code="39259090", pieces_per_box=prod.get("ppb", 1), std_packaging=prod.get("ppb", 1))
            db.add(p)
            db.flush()
            db.add(Pricing(product_id=p.id, raw_material_cost=prod["rate"], total_cost=prod["rate"], profit_margin=20, gst_rate=18, mrp=prod["mrp"]))
            pid += 1

        db.commit()
        print(f"Seeded {pid - 1} products")
    finally:
        db.close()


class ProductIn(BaseModel):
    part_no: str = ""
    name: str
    category: str = ""
    size: str = ""
    load_rating: str = ""
    material: str = "FRP"
    color: str = "Grey"
    unit: str = "Nos"
    hsn_code: str = ""


class PricingIn(BaseModel):
    raw_material_cost: float = 0
    mrp: float = 0
    labor_cost: float = 0
    overhead_cost: float = 0
    packing_cost: float = 0
    profit_margin: float = 20
    gst_rate: float = 18


class CustomerIn(BaseModel):
    customer_id: str
    gstin: str = ""
    billing_address: str = ""
    shipping_address: str = ""
    state: str = ""
    district: str = ""
    city: str = ""
    pincode: str = ""
    contact_name: str = ""
    contact_number: str = ""
    contact_email: str = ""
    exec_code: str = ""
    exec_name: str = ""
    exec_number: str = ""
    exec_email: str = ""
    blacklisted: int = 0


class TransporterIn(BaseModel):
    transporter_id: str
    name: str
    phone: str
    email: str
    address: str
    state: str
    district: str
    city: str
    pincode: str
    gst_number: str
    pan_number: str
    gst_certificate: str = ""
    pan_card: str = ""
    contact_person: str
    contact_number: str
    blacklisted: int = 0


class SaleIn(BaseModel):
    customer_id: int
    product_id: int
    quantity: int
    unit_price: float
    discount_percent: float = 0
    freight_amount: float = 0
    invoice_value: float = 0
    payment_status: str = "Pending"
    payment_method: str = "Cash"
    notes: str = ""


class ExpenseIn(BaseModel):
    category: str
    description: str = ""
    amount: float
    vendor: str = ""
    expense_date: Optional[str] = None


class OrderIn(BaseModel):
    sl_no: int = 0
    po_no: str = ""
    po_date: str = ""
    customer_name: str = ""
    billing_site: str = ""
    shipping_site: str = ""
    no_of_boxes: int = 0
    value_excl_gst_freight: float = 0
    invoice_no: str = ""
    invoice_date: str = ""
    invoice_amount_excl_gst: float = 0
    weight_kgs: float = 0
    freight_rate_per_kg: float = 0
    transport_charges: float = 0
    invoice_amount: float = 0
    eway_bill_no: str = ""
    lr_no: str = ""
    entry_date: str = ""
    credit_note_amount: float = 0
    credit_note_no: str = ""
    transporter: str = ""
    transporter_no: str = ""


class ProformaOrderItemIn(BaseModel):
    product_id: int
    part_no: str = ""
    description: str = ""
    size: str = ""
    category: str = ""
    qty_boxes: int = 1
    std_packaging: int = 1
    pieces_per_box: int = 1
    final_qty: int = 0
    mrp: float = 0
    d1: float = 0
    d2: float = 0
    d3: float = 0
    d4: float = 0
    d5: float = 0
    cd: float = 0
    discount_percent: float = 0
    net_rate: float = 0
    lock_hinge: int = 0
    basic_amount: float = 0


class ProformaOrderIn(BaseModel):
    customer_id: int
    billing_site: str = ""
    shipping_site: str = ""
    freight_amount: float = 0
    payment_status: str = "Pending"
    payment_method: str = "Cash"
    transport_mode: str = ""
    delivery_days: int = 30
    notes: str = ""
    terms: str = ""
    order_type: str = "PI"
    items: List[ProformaOrderItemIn] = []


# ---- PRODUCTS ----
SIZE_ORDER = {
    "10x10": 1, "250x250": 1,
    "12x12": 2, "300x300": 2,
    "15x15": 3, "380x380": 3,
    "18x18": 4, "450x450": 4,
    "21x21": 5, "535x535": 5,
    "24x24": 6, "600x600": 6,
    "26x26": 7, "660x660": 7,
    "28x28": 8, "710x710": 8,
    "30x30": 9, "760x760": 9,
    "36x36": 10, "900x900": 10,
    "42x42": 11, "1065x1065": 11,
    "12x18": 12, "300x450": 12,
    "12x24": 13, "300x600": 13,
    "18x24": 14, "450x600": 14,
}
COLOR_ORDER = {"Grey": 0, "White": 1}
CATEGORY_ORDER = {"Manhole Cover": 0, "Gully Cover": 1}

CSV_ORDER = {pn: i for i, (pn, _) in enumerate(PART_NO_CSV)}

@app.get("/api/products")
def list_products():
    db = SessionLocal()
    try:
        rows = db.query(Product).all()
        out = []
        for p in rows:
            mrp = p.pricing.mrp if p.pricing else 0
            out.append({
                "id": p.id, "part_no": p.part_no, "name": p.name, "category": p.category,
                "size": p.size, "load_rating": p.load_rating,
                "material": p.material, "color": p.color, "unit": p.unit, "hsn_code": p.hsn_code,
                "mrp": mrp
            })
        out.sort(key=lambda p: CSV_ORDER.get(p["part_no"], 999))
        return out
    finally:
        db.close()


@app.post("/api/products")
def create_product(inp: ProductIn):
    db = SessionLocal()
    try:
        p = Product(**inp.dict())
        db.add(p)
        db.commit()
        db.refresh(p)
        db.add(Pricing(product_id=p.id))
        db.commit()
        return {"id": p.id, "message": "Product created"}
    finally:
        db.close()


@app.put("/api/products/{pid}")
def update_product(pid: int, inp: ProductIn):
    db = SessionLocal()
    try:
        p = db.query(Product).filter(Product.id == pid).first()
        if not p:
            raise HTTPException(404, "Not found")
        for k, v in inp.dict().items():
            setattr(p, k, v)
        db.commit()
        return {"message": "Updated"}
    finally:
        db.close()


@app.delete("/api/products/{pid}")
def delete_product(pid: int):
    db = SessionLocal()
    try:
        p = db.query(Product).filter(Product.id == pid).first()
        if not p:
            raise HTTPException(404, "Not found")
        db.delete(p)
        db.commit()
        return {"message": "Deleted"}
    finally:
        db.close()


# ---- PRICING ----
@app.get("/api/products/{pid}/pricing")
def get_pricing(pid: int):
    db = SessionLocal()
    try:
        pr = db.query(Pricing).filter(Pricing.product_id == pid).first()
        if not pr:
            raise HTTPException(404, "Not found")
        return {
            "raw_material_cost": pr.raw_material_cost,
            "labor_cost": pr.labor_cost,
            "overhead_cost": pr.overhead_cost,
            "packing_cost": pr.packing_cost,
            "total_cost": pr.total_cost,
            "profit_margin": pr.profit_margin,
            "mrp": pr.mrp,
            "dealer_price": pr.dealer_price,
            "distributor_price": pr.distributor_price,
            "gst_rate": pr.gst_rate
        }
    finally:
        db.close()


@app.put("/api/products/{pid}/pricing")
def update_pricing(pid: int, inp: PricingIn):
    db = SessionLocal()
    try:
        pr = db.query(Pricing).filter(Pricing.product_id == pid).first()
        if not pr:
            pr = Pricing(product_id=pid)
            db.add(pr)
        pr.raw_material_cost = inp.raw_material_cost
        pr.mrp = inp.mrp
        pr.gst_rate = inp.gst_rate
        db.commit()
        return {"message": "Updated", "mrp": inp.mrp}
    finally:
        db.close()


# ---- ORDERS ----
@app.get("/api/orders")
def list_orders():
    db = SessionLocal()
    try:
        rows = db.query(Order).order_by(Order.id).all()
        return [{"id": o.id, "sl_no": o.sl_no, "po_no": o.po_no, "po_date": o.po_date,
                 "customer_name": o.customer_name or "",
                 "billing_site": o.billing_site, "shipping_site": o.shipping_site,
                 "no_of_boxes": o.no_of_boxes, "value_excl_gst_freight": o.value_excl_gst_freight,
                 "invoice_no": o.invoice_no, "invoice_date": o.invoice_date,
                 "invoice_amount_excl_gst": o.invoice_amount_excl_gst,
                 "weight_kgs": o.weight_kgs, "freight_rate_per_kg": o.freight_rate_per_kg,
                 "transport_charges": o.transport_charges, "invoice_amount": o.invoice_amount,
                 "eway_bill_no": o.eway_bill_no, "lr_no": o.lr_no, "entry_date": o.entry_date,
                 "credit_note_amount": o.credit_note_amount, "credit_note_no": o.credit_note_no,
                 "transporter": o.transporter, "transporter_no": o.transporter_no}
                for o in rows]
    finally:
        db.close()


@app.post("/api/orders")
def create_order(inp: OrderIn):
    db = SessionLocal()
    try:
        max_sl = db.query(Order.sl_no).order_by(Order.sl_no.desc()).first()
        next_sl = (max_sl[0] + 1) if max_sl and max_sl[0] else 1
        data = inp.dict()
        data["sl_no"] = next_sl
        o = Order(**data)
        db.add(o)
        db.commit()
        db.refresh(o)
        return {"id": o.id, "sl_no": next_sl, "message": "Order created"}
    finally:
        db.close()


@app.put("/api/orders/{oid}")
def update_order(oid: int, inp: OrderIn):
    db = SessionLocal()
    try:
        o = db.query(Order).filter(Order.id == oid).first()
        if not o:
            raise HTTPException(404, "Not found")
        for k, v in inp.dict().items():
            setattr(o, k, v)
        db.commit()
        return {"message": "Updated"}
    finally:
        db.close()


@app.delete("/api/orders/{oid}")
def delete_order(oid: int):
    db = SessionLocal()
    try:
        o = db.query(Order).filter(Order.id == oid).first()
        if not o:
            raise HTTPException(404, "Not found")
        db.delete(o)
        db.commit()
        return {"message": "Deleted"}
    finally:
        db.close()


# ---- PROFORMA ORDERS (Multi-Product PI/PO) ----
@app.get("/api/proforma-orders")
def list_proforma_orders(order_type: str = None):
    db = SessionLocal()
    try:
        query = db.query(ProformaOrder)
        if order_type:
            query = query.filter(ProformaOrder.order_type == order_type)
        rows = query.order_by(ProformaOrder.created_at.desc()).all()
        out = []
        for o in rows:
            cust = db.query(Customer).filter(Customer.id == o.customer_id).first()
            items = db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id == o.id).all()
            out.append({
                "id": o.id, "pi_no": o.pi_no,
                "pi_date": o.pi_date.isoformat() if o.pi_date else None,
                "customer_name": cust.contact_name if cust else "?",
                "customer_id": o.customer_id,
                "billing_site": o.billing_site, "shipping_site": o.shipping_site,
                "no_of_boxes": o.no_of_boxes, "total_qty": o.total_qty,
                "value_excl_gst": o.value_excl_gst, "gst_amount": o.gst_amount,
                "total_amount": o.total_amount, "freight_amount": o.freight_amount,
                "payment_status": o.payment_status, "payment_method": o.payment_method,
                "transport_mode": o.transport_mode, "delivery_days": o.delivery_days,
                "notes": o.notes, "terms": o.terms, "order_type": o.order_type,
                "item_count": len(items),
                "created_at": o.created_at.isoformat() if o.created_at else None
            })
        return out
    finally:
        db.close()


@app.get("/api/proforma-orders/{oid}")
def get_proforma_order(oid: int):
    db = SessionLocal()
    try:
        o = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not o:
            raise HTTPException(404, "Order not found")
        cust = db.query(Customer).filter(Customer.id == o.customer_id).first()
        items = db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id == o.id).order_by(ProformaOrderItem.sl_no).all()
        items_out = []
        for item in items:
            prod = db.query(Product).filter(Product.id == item.product_id).first()
            items_out.append({
                "id": item.id, "sl_no": item.sl_no,
                "product_id": item.product_id,
                "product_name": prod.name if prod else "?",
                "part_no": item.part_no, "description": item.description,
                "size": item.size, "category": item.category,
                "qty_boxes": item.qty_boxes, "std_packaging": item.std_packaging,
                "pieces_per_box": item.pieces_per_box, "final_qty": item.final_qty,
                "mrp": item.mrp, "d1": item.d1, "d2": item.d2, "d3": item.d3,
                "d4": item.d4, "d5": item.d5, "cd": item.cd,
                "discount_percent": item.discount_percent,
                "net_rate": item.net_rate, "lock_hinge": item.lock_hinge,
                "basic_amount": item.basic_amount
            })
        return {
            "id": o.id, "pi_no": o.pi_no,
            "pi_date": o.pi_date.isoformat() if o.pi_date else None,
            "customer_id": o.customer_id,
            "customer_name": cust.contact_name if cust else "?",
            "customer_gst": cust.gstin if cust else "",
            "customer_state": cust.state if cust else "",
            "customer_address": cust.billing_address if cust else "",
            "billing_site": o.billing_site, "shipping_site": o.shipping_site,
            "no_of_boxes": o.no_of_boxes, "total_qty": o.total_qty,
            "value_excl_gst": o.value_excl_gst, "gst_amount": o.gst_amount,
            "total_amount": o.total_amount, "freight_amount": o.freight_amount,
            "payment_status": o.payment_status, "payment_method": o.payment_method,
            "transport_mode": o.transport_mode, "delivery_days": o.delivery_days,
            "notes": o.notes, "terms": o.terms, "order_type": o.order_type,
            "items": items_out,
            "created_at": o.created_at.isoformat() if o.created_at else None
        }
    finally:
        db.close()


@app.post("/api/proforma-orders")
def create_proforma_order(inp: ProformaOrderIn):
    db = SessionLocal()
    try:
        customer = db.query(Customer).filter(Customer.id == inp.customer_id).first()
        if not customer:
            raise HTTPException(404, "Customer not found")

        max_id = db.query(ProformaOrder).count()
        pi_no = f"RFC/{datetime.now().strftime('%y%m')}-{max_id + 1:03d}"

        total_qty = 0
        total_basic = 0

        for item in inp.items:
            total_qty += item.final_qty
            total_basic += item.basic_amount

        gst_amount = total_basic * 0.18
        total_amount = total_basic + gst_amount + inp.freight_amount

        order = ProformaOrder(
            pi_no=pi_no, customer_id=inp.customer_id,
            billing_site=inp.billing_site, shipping_site=inp.shipping_site,
            total_qty=total_qty, no_of_boxes=sum(i.qty_boxes for i in inp.items),
            value_excl_gst=total_basic, gst_amount=gst_amount,
            total_amount=total_amount, freight_amount=inp.freight_amount,
            payment_status=inp.payment_status, payment_method=inp.payment_method,
            transport_mode=inp.transport_mode, delivery_days=inp.delivery_days,
            notes=inp.notes, terms=inp.terms, order_type=inp.order_type
        )
        db.add(order)
        db.flush()

        for idx, item in enumerate(inp.items):
            db.add(ProformaOrderItem(
                proforma_order_id=order.id, sl_no=idx + 1,
                product_id=item.product_id, part_no=item.part_no,
                description=item.description, size=item.size,
                category=item.category, qty_boxes=item.qty_boxes,
                std_packaging=item.std_packaging, pieces_per_box=item.pieces_per_box,
                final_qty=item.final_qty, mrp=item.mrp,
                d1=item.d1, d2=item.d2, d3=item.d3, d4=item.d4, d5=item.d5,
                cd=item.cd, discount_percent=item.discount_percent,
                net_rate=item.net_rate, lock_hinge=item.lock_hinge,
                basic_amount=item.basic_amount
            ))

        db.commit()
        return {"id": order.id, "pi_no": pi_no, "total": total_amount}
    finally:
        db.close()


@app.put("/api/proforma-orders/{oid}")
def update_proforma_order(oid: int, inp: ProformaOrderIn):
    db = SessionLocal()
    try:
        order = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not order:
            raise HTTPException(404, "Order not found")

        customer = db.query(Customer).filter(Customer.id == inp.customer_id).first()
        if not customer:
            raise HTTPException(404, "Customer not found")

        order.customer_id = inp.customer_id
        order.billing_site = inp.billing_site
        order.shipping_site = inp.shipping_site
        order.freight_amount = inp.freight_amount
        order.payment_status = inp.payment_status
        order.payment_method = inp.payment_method
        order.transport_mode = inp.transport_mode
        order.delivery_days = inp.delivery_days
        order.notes = inp.notes
        order.terms = inp.terms
        order.order_type = inp.order_type

        db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id == oid).delete()

        total_qty = 0
        total_basic = 0
        for idx, item in enumerate(inp.items):
            total_qty += item.final_qty
            total_basic += item.basic_amount
            db.add(ProformaOrderItem(
                proforma_order_id=oid, sl_no=idx + 1,
                product_id=item.product_id, part_no=item.part_no,
                description=item.description, size=item.size,
                category=item.category, qty_boxes=item.qty_boxes,
                std_packaging=item.std_packaging, pieces_per_box=item.pieces_per_box,
                final_qty=item.final_qty, mrp=item.mrp,
                d1=item.d1, d2=item.d2, d3=item.d3, d4=item.d4, d5=item.d5,
                cd=item.cd, discount_percent=item.discount_percent,
                net_rate=item.net_rate, lock_hinge=item.lock_hinge,
                basic_amount=item.basic_amount
            ))

        gst_amount = total_basic * 0.18
        order.total_qty = total_qty
        order.no_of_boxes = sum(i.qty_boxes for i in inp.items)
        order.value_excl_gst = total_basic
        order.gst_amount = gst_amount
        order.total_amount = total_basic + gst_amount + inp.freight_amount
        order.updated_at = datetime.utcnow()

        db.commit()
        return {"message": "Order updated"}
    finally:
        db.close()


@app.delete("/api/proforma-orders/{oid}")
def delete_proforma_order(oid: int):
    db = SessionLocal()
    try:
        order = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not order:
            raise HTTPException(404, "Order not found")
        db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id == oid).delete()
        db.delete(order)
        db.commit()
        return {"message": "Order deleted"}
    finally:
        db.close()


@app.get("/api/products/{pid}/details")
def get_product_details(pid: int):
    db = SessionLocal()
    try:
        p = db.query(Product).filter(Product.id == pid).first()
        if not p:
            raise HTTPException(404, "Product not found")
        pr = db.query(Pricing).filter(Pricing.product_id == pid).first()
        return {
            "id": p.id, "name": p.name, "category": p.category,
            "size": p.size, "part_no": p.part_no,
            "mrp": pr.mrp if pr else 0,
            "pieces_per_box": p.pieces_per_box or 1,
            "std_packaging": p.std_packaging or 1
        }
    finally:
        db.close()


@app.get("/api/proforma-orders/{oid}/pdf")
def generate_proforma_order_pdf(oid: int):
    db = SessionLocal()
    try:
        order = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not order:
            raise HTTPException(404, "Order not found")
        customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
        items = db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id == oid).order_by(ProformaOrderItem.sl_no).all()

        items_html = ""
        for item in items:
            items_html += f"""
            <tr>
                <td style="padding:6px;border:1px solid #ddd;text-align:center;">{item.sl_no}</td>
                <td style="padding:6px;border:1px solid #ddd;">{item.description or ''} ({item.size or ''})</td>
                <td style="padding:6px;border:1px solid #ddd;">{item.category or '-'}</td>
                <td style="padding:6px;border:1px solid #ddd;">{item.part_no or '-'}</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:center;">{item.qty_boxes}</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:center;">Boxes</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:center;">{item.std_packaging}</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:center;">{item.final_qty}</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:center;">Pieces</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:right;">₹{item.mrp:,.2f}</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:center;">{item.discount_percent}%</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:right;">₹{item.net_rate:,.2f}</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:center;">{item.lock_hinge}</td>
                <td style="padding:6px;border:1px solid #ddd;text-align:right;font-weight:bold;">₹{item.basic_amount:,.2f}</td>
            </tr>"""

        title = "PROFORMA INVOICE" if order.order_type == "PI" else "PURCHASE ORDER"
        pi_date = order.pi_date.strftime("%d-%b-%Y") if order.pi_date else ""

        html = f"""<!DOCTYPE html>
<html><head><title>{title}</title>
<style>
body{{font-family:Arial,sans-serif;margin:20px;font-size:12px;}}
table{{width:100%;border-collapse:collapse;}}
.header{{text-align:center;margin-bottom:20px;}}
.header h1{{margin:0;font-size:18px;color:#1a365d;}}
.header p{{margin:2px 0;color:#555;font-size:11px;}}
.details{{margin:15px 0;}}
.details td{{padding:3px 8px;font-size:11px;}}
.totals{{text-align:right;margin-top:10px;}}
.totals td{{padding:4px 8px;font-size:11px;}}
.terms{{margin-top:20px;font-size:10px;border-top:1px solid #ccc;padding-top:10px;}}
@media print{{body{{margin:10mm;}}}}
</style></head><body>

<div class="header">
<h1>RANA FORGING PVT LTD</h1>
<p>KHATU MOHAMADPUR, NH-8, JAIPUR-302012 (RAJ.)</p>
<p>Tel: 9928684835, 9672962255 | Email: ranaforging@gmail.com</p>
<p>GSTIN: 08AAFCT3014D1ZC | PAN: AAFCT3014D | State: RAJASTHAN (08)</p>
</div>

<div class="details"><table>
<tr><td style="font-weight:bold;width:120px;">{"Quotation No:" if order.order_type == "PI" else "PO No:"}</td><td>{order.pi_no}</td>
<td style="font-weight:bold;width:80px;">Date:</td><td>{pi_date}</td></tr>
<tr><td style="font-weight:bold;">Customer:</td><td colspan="3">{customer.contact_name if customer else "?"}</td></tr>
{"<tr><td style='font-weight:bold;'>GSTIN:</td><td colspan='3'>" + (customer.gstin or "") + "</td></tr>" if customer and customer.gstin else ""}
{"<tr><td style='font-weight:bold;'>Billing Site:</td><td colspan='3'>" + (order.billing_site or "") + "</td></tr>" if order.billing_site else ""}
{"<tr><td style='font-weight:bold;'>Shipping Site:</td><td colspan='3'>" + (order.shipping_site or "") + "</td></tr>" if order.shipping_site else ""}
</table></div>

<table style="margin-top:15px;">
<thead><tr style="background:#1a365d;color:white;">
<th style="padding:6px;border:1px solid #ddd;">Sr.</th>
<th style="padding:6px;border:1px solid #ddd;">Description</th>
<th style="padding:6px;border:1px solid #ddd;">Item Group</th>
<th style="padding:6px;border:1px solid #ddd;">Part No</th>
<th style="padding:6px;border:1px solid #ddd;">Boxes</th>
<th style="padding:6px;border:1px solid #ddd;">Base UOM</th>
<th style="padding:6px;border:1px solid #ddd;">Std Pkg</th>
<th style="padding:6px;border:1px solid #ddd;">Final Qty</th>
<th style="padding:6px;border:1px solid #ddd;">Final UOM</th>
<th style="padding:6px;border:1px solid #ddd;">MRP</th>
<th style="padding:6px;border:1px solid #ddd;">Disc %</th>
<th style="padding:6px;border:1px solid #ddd;">Net Rate</th>
<th style="padding:6px;border:1px solid #ddd;">L&H</th>
<th style="padding:6px;border:1px solid #ddd;">Basic Amt</th>
</tr></thead><tbody>{items_html}</tbody></table>

<div class="totals"><table style="width:350px;margin-left:auto;">
<tr><td>Sub Total:</td><td style="text-align:right;">₹{order.value_excl_gst:,.2f}</td></tr>
<tr><td>GST @18%:</td><td style="text-align:right;">₹{order.gst_amount:,.2f}</td></tr>
<tr><td>Freight:</td><td style="text-align:right;">₹{order.freight_amount:,.2f}</td></tr>
<tr style="font-size:14px;font-weight:bold;border-top:2px solid #1a365d;"><td>Grand Total:</td><td style="text-align:right;color:#16a34a;">₹{order.total_amount:,.2f}</td></tr>
</table></div>

<div class="terms"><h4>Terms & Conditions:</h4><ol>
<li>GST @ 18% will be charged extra</li>
<li>Freight will be charged extra</li>
<li>Material will be supplied ex-factory</li>
<li>Payment: Advance Cheque/Draft</li>
<li>Delivery: {order.delivery_days} Days</li>
<li>Our Rc. No. is to be quoted on your invoice</li>
<li>Insurance: To be arranged by the buyer</li>
<li>Packing: Standard packaging</li>
<li>Subject to Jaipur Jurisdiction only</li>
</ol></div>

<div style="margin-top:40px;text-align:right;"><p>For RANA FORGING PVT LTD</p>
<p style="margin-top:30px;">Authorized Signatory</p></div>

</body></html>"""

        return HTMLResponse(content=html)
    finally:
        db.close()


# ---- CUSTOMERS ----
@app.get("/api/customers")
def list_customers():
    db = SessionLocal()
    try:
        rows = db.query(Customer).all()
        return [{"id": c.id, "customer_id": c.customer_id, "gstin": c.gstin,
                 "billing_address": c.billing_address, "shipping_address": c.shipping_address,
                 "state": c.state, "district": c.district, "city": c.city, "pincode": c.pincode,
                 "contact_name": c.contact_name, "contact_number": c.contact_number, "contact_email": c.contact_email,
                 "exec_code": c.exec_code, "exec_name": c.exec_name, "exec_number": c.exec_number, "exec_email": c.exec_email,
                 "blacklisted": c.blacklisted}
                for c in rows]
    finally:
        db.close()


@app.post("/api/customers")
def create_customer(inp: CustomerIn):
    db = SessionLocal()
    try:
        c = Customer(**inp.dict())
        db.add(c)
        db.commit()
        db.refresh(c)
        return {"id": c.id, "message": "Customer created"}
    finally:
        db.close()


@app.put("/api/customers/{cid}")
def update_customer(cid: int, inp: CustomerIn):
    db = SessionLocal()
    try:
        c = db.query(Customer).filter(Customer.id == cid).first()
        if not c:
            raise HTTPException(404, "Not found")
        for k, v in inp.dict().items():
            setattr(c, k, v)
        db.commit()
        return {"message": "Updated"}
    finally:
        db.close()


@app.delete("/api/customers/{cid}")
def delete_customer(cid: int):
    db = SessionLocal()
    try:
        c = db.query(Customer).filter(Customer.id == cid).first()
        if not c:
            raise HTTPException(404, "Not found")
        db.delete(c)
        db.commit()
        return {"message": "Deleted"}
    finally:
        db.close()


# ---- TRANSPORTERS ----
@app.get("/api/transporters")
def list_transporters():
    db = SessionLocal()
    try:
        rows = db.query(Transporter).all()
        return [{"id": t.id, "transporter_id": t.transporter_id, "name": t.name,
                 "phone": t.phone, "email": t.email, "address": t.address,
                 "state": t.state, "district": t.district, "city": t.city, "pincode": t.pincode,
                 "gst_number": t.gst_number, "pan_number": t.pan_number,
                 "gst_certificate": t.gst_certificate, "pan_card": t.pan_card,
                 "contact_person": t.contact_person, "contact_number": t.contact_number,
                 "blacklisted": t.blacklisted}
                for t in rows]
    finally:
        db.close()


@app.post("/api/transporters")
def create_transporter(inp: TransporterIn):
    db = SessionLocal()
    try:
        t = Transporter(**inp.dict())
        db.add(t)
        db.commit()
        db.refresh(t)
        return {"id": t.id, "message": "Transporter created"}
    finally:
        db.close()


@app.put("/api/transporters/{tid}")
def update_transporter(tid: int, inp: TransporterIn):
    db = SessionLocal()
    try:
        t = db.query(Transporter).filter(Transporter.id == tid).first()
        if not t:
            raise HTTPException(404, "Not found")
        for k, v in inp.dict().items():
            setattr(t, k, v)
        db.commit()
        return {"message": "Updated"}
    finally:
        db.close()


@app.get("/api/fix-urls")
def fix_urls():
    db = SessionLocal()
    try:
        rows = db.query(Transporter).all()
        fixed = 0
        for t in rows:
            changed = False
            for field in ["gst_certificate", "pan_card"]:
                url = getattr(t, field, "")
                if url and "/image/upload/" in url and url.endswith(".pdf"):
                    setattr(t, field, url.replace("/image/upload/", "/raw/upload/"))
                    changed = True
            if changed:
                fixed += 1
        db.commit()
        return {"fixed": fixed}
    finally:
        db.close()


@app.get("/api/view-file")
async def view_file(url: str = Query(...)):
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req) as resp:
            data = resp.read()
            if data[:4] == b'%PDF':
                content_type = 'application/pdf'
            elif data[:2] == b'\xff\xd8':
                content_type = 'image/jpeg'
            elif data[:8] == b'\x89PNG\r\n\x1a\n':
                content_type = 'image/png'
            else:
                content_type = resp.headers.get("Content-Type", "application/octet-stream")
            return Response(content=data, media_type=content_type, headers={
                "Content-Disposition": "inline",
            })
    except Exception as e:
        raise HTTPException(500, f"Failed to load file: {str(e)}")


@app.delete("/api/transporters/{tid}")
def delete_transporter(tid: int):
    db = SessionLocal()
    try:
        t = db.query(Transporter).filter(Transporter.id == tid).first()
        if not t:
            raise HTTPException(404, "Not found")
        db.delete(t)
        db.commit()
        return {"message": "Deleted"}
    finally:
        db.close()


# ---- SALES ----
@app.get("/api/sales")
def list_sales():
    db = SessionLocal()
    try:
        rows = db.query(Sale).order_by(Sale.id.desc()).all()
        out = []
        for s in rows:
            try:
                cust_name = ""
                if s.customer_id:
                    cust = db.query(Customer).filter(Customer.id == s.customer_id).first()
                    cust_name = cust.contact_name if cust else ""
                party = s.party_name or cust_name or ""
                loc = s.location or ""
                out.append({
                    "id": s.id, "invoice_no": s.invoice_no or "",
                    "customer_id": s.customer_id or None,
                    "product_id": s.product_id or None,
                    "quantity": s.quantity or 0,
                    "unit_price": s.unit_price or 0,
                    "discount_percent": s.discount_percent or 0,
                    "freight_amount": s.freight_amount or 0,
                    "payment_status": s.payment_status or "",
                    "payment_method": s.payment_method or "",
                    "notes": s.notes or "",
                    "party_name": party, "location": loc,
                    "state": s.state or "",
                    "transporter_name": s.transporter_name or "",
                    "lr_no": s.lr_no or "",
                    "weight_kgs": s.weight_kgs or 0,
                    "gp": s.gp or 0,
                    "gp_percent": s.gp_percent or 0,
                    "invoice_value": s.invoice_value or 0,
                    "total_amount": s.total_amount or 0,
                    "sale_date": s.sale_date.isoformat() if s.sale_date else None,
                    "payment_terms": s.payment_terms or "",
                    "source_csv": s.source_csv or "",
                })
            except Exception:
                continue
        return out
    finally:
        db.close()


@app.post("/api/sales")
def create_sale(inp: SaleIn):
    db = SessionLocal()
    try:
        prod = db.query(Product).filter(Product.id == inp.product_id).first()
        if not prod:
            raise HTTPException(404, "Product not found")
        pr = db.query(Pricing).filter(Pricing.product_id == inp.product_id).first()
        gst_rate = pr.gst_rate if pr else 18

        taxable = inp.quantity * inp.unit_price
        disc_amt = taxable * inp.discount_percent / 100
        taxable -= disc_amt
        cgst = taxable * gst_rate / 200
        sgst = taxable * gst_rate / 200
        total = taxable + cgst + sgst + inp.freight_amount

        max_id = db.query(Sale).count()
        invoice_no = f"RFRP-{max_id + 1:05d}"

        s = Sale(
            invoice_no=invoice_no, customer_id=inp.customer_id,
            product_id=inp.product_id, quantity=inp.quantity,
            unit_price=inp.unit_price, discount_percent=inp.discount_percent,
            discount_amount=disc_amt, taxable_amount=taxable,
            cgst_rate=gst_rate / 2, cgst_amount=cgst,
            sgst_rate=gst_rate / 2, sgst_amount=sgst,
            freight_amount=inp.freight_amount, total_amount=total,
            payment_status=inp.payment_status, payment_method=inp.payment_method,
            invoice_value=inp.invoice_value, notes=inp.notes
        )
        db.add(s)
        db.commit()
        return {"invoice_no": invoice_no, "total": total}
    finally:
        db.close()


@app.delete("/api/sales/{sid}")
def delete_sale(sid: int):
    db = SessionLocal()
    try:
        s = db.query(Sale).filter(Sale.id == sid).first()
        if not s:
            raise HTTPException(404, "Not found")
        db.delete(s)
        db.commit()
        return {"message": "Deleted"}
    finally:
        db.close()


@app.put("/api/sales/{sid}")
def update_sale(sid: int, inp: SaleIn):
    db = SessionLocal()
    try:
        s = db.query(Sale).filter(Sale.id == sid).first()
        if not s:
            raise HTTPException(404, "Not found")

        prod = db.query(Product).filter(Product.id == inp.product_id).first()
        if not prod:
            raise HTTPException(404, "Product not found")
        pr = db.query(Pricing).filter(Pricing.product_id == inp.product_id).first()
        gst_rate = pr.gst_rate if pr else 18

        taxable = inp.quantity * inp.unit_price
        disc_amt = taxable * inp.discount_percent / 100
        taxable -= disc_amt
        cgst = taxable * gst_rate / 200
        sgst = taxable * gst_rate / 200
        total = taxable + cgst + sgst + inp.freight_amount

        s.customer_id = inp.customer_id
        s.product_id = inp.product_id
        s.quantity = inp.quantity
        s.unit_price = inp.unit_price
        s.discount_percent = inp.discount_percent
        s.discount_amount = disc_amt
        s.taxable_amount = taxable
        s.cgst_rate = gst_rate / 2
        s.cgst_amount = cgst
        s.sgst_rate = gst_rate / 2
        s.sgst_amount = sgst
        s.freight_amount = inp.freight_amount
        s.total_amount = total
        s.payment_status = inp.payment_status
        s.payment_method = inp.payment_method
        s.invoice_value = inp.invoice_value
        s.notes = inp.notes

        db.commit()
        return {"message": "Sale updated", "total": total}
    finally:
        db.close()


# ---- EXPENSES ----
@app.get("/api/expenses")
def list_expenses():
    db = SessionLocal()
    try:
        rows = db.query(Expense).order_by(Expense.expense_date.desc()).all()
        return [{"id": e.id, "category": e.category, "description": e.description,
                 "amount": e.amount, "vendor": e.vendor,
                 "expense_date": e.expense_date.isoformat() if e.expense_date else None}
                for e in rows]
    finally:
        db.close()


@app.post("/api/expenses")
def create_expense(inp: ExpenseIn):
    db = SessionLocal()
    try:
        dt = datetime.strptime(inp.expense_date, "%Y-%m-%d") if inp.expense_date else datetime.utcnow()
        e = Expense(category=inp.category, description=inp.description,
                    amount=inp.amount, vendor=inp.vendor, expense_date=dt)
        db.add(e)
        db.commit()
        return {"message": "Expense added"}
    finally:
        db.close()


@app.delete("/api/expenses/{eid}")
def delete_expense(eid: int):
    db = SessionLocal()
    try:
        e = db.query(Expense).filter(Expense.id == eid).first()
        if not e:
            raise HTTPException(404, "Not found")
        db.delete(e)
        db.commit()
        return {"message": "Deleted"}
    finally:
        db.close()


@app.put("/api/expenses/{eid}")
def update_expense(eid: int, inp: ExpenseIn):
    db = SessionLocal()
    try:
        e = db.query(Expense).filter(Expense.id == eid).first()
        if not e:
            raise HTTPException(404, "Not found")
        e.category = inp.category
        e.description = inp.description
        e.amount = inp.amount
        e.vendor = inp.vendor
        if inp.expense_date:
            try:
                e.expense_date = datetime.strptime(inp.expense_date, "%Y-%m-%d")
            except Exception:
                pass
        db.commit()
        return {"message": "Expense updated"}
    finally:
        db.close()


# ---- REPORTS ----
@app.get("/api/reports/profit-loss")
def profit_loss(start_date: str = None, end_date: str = None):
    db = SessionLocal()
    try:
        sales_q = db.query(Sale)
        expenses_q = db.query(Expense)
        orders_q = db.query(Order)

        if start_date:
            try:
                sd = datetime.strptime(start_date, "%Y-%m-%d")
                sales_q = sales_q.filter(Sale.sale_date >= sd)
                expenses_q = expenses_q.filter(Expense.expense_date >= sd)
                orders_q = orders_q.filter(Order.entry_date >= start_date)
            except Exception:
                pass
        if end_date:
            try:
                ed = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                sales_q = sales_q.filter(Sale.sale_date <= ed)
                expenses_q = expenses_q.filter(Expense.expense_date <= ed)
                orders_q = orders_q.filter(Order.entry_date <= end_date)
            except Exception:
                pass

        sales = sales_q.all()
        expenses = expenses_q.all()
        orders = orders_q.all()

        # Revenue from Sales
        sale_revenue = sum(s.total_amount or 0 for s in sales)
        sale_freight = sum(s.freight_amount or 0 for s in sales)
        units = sum(s.quantity or 0 for s in sales)
        gst = sum((s.cgst_amount or 0) + (s.sgst_amount or 0) for s in sales)
        gp_from_sales = sum(s.gp or 0 for s in sales)
        gp_values = [s.gp_percent for s in sales if s.gp_percent and s.gp_percent > 0]
        gp_avg = sum(gp_values) / len(gp_values) if gp_values else 0

        # COGS from Orders (cost of goods)
        order_cogs = sum(o.value_excl_gst_freight or 0 for o in orders)
        order_freight_cost = sum(o.transport_charges or 0 for o in orders)
        order_invoice_total = sum(o.invoice_amount or 0 for o in orders)
        order_boxes = sum(o.no_of_boxes or 0 for o in orders)
        order_weight = sum(o.weight_kgs or 0 for o in orders)
        order_credit_notes = sum(o.credit_note_amount or 0 for o in orders)

        # Total revenue (sale revenue + freight from sales)
        total_revenue = sale_revenue + sale_freight

        # Total COGS (order cost + transport)
        total_cogs = order_cogs + order_freight_cost

        exp_by_cat = {}
        for e in expenses:
            exp_by_cat[e.category] = exp_by_cat.get(e.category, 0) + e.amount
        total_opex = sum(exp_by_cat.values())

        gross_profit = total_revenue - total_cogs - order_credit_notes
        if gp_from_sales > 0:
            gross_profit = gp_from_sales
        gross_margin = (gross_profit / total_revenue * 100) if total_revenue else 0

        ebitda = gross_profit - total_opex
        tax = ebitda * 0.25 if ebitda > 0 else 0
        pat = ebitda - tax

        return {
            "total_revenue": total_revenue,
            "sale_revenue": sale_revenue, "sale_freight": sale_freight,
            "total_cogs": total_cogs,
            "order_cogs": order_cogs, "order_freight_cost": order_freight_cost,
            "order_invoice_total": order_invoice_total,
            "order_boxes": order_boxes, "order_weight": order_weight,
            "order_credit_notes": order_credit_notes,
            "gst": gst, "units": units,
            "gross_profit": gross_profit, "gross_margin": gross_margin,
            "gp_from_sales": gp_from_sales, "gp_avg": gp_avg,
            "expenses": exp_by_cat, "total_opex": total_opex,
            "ebitda": ebitda, "ebitda_margin": (ebitda / total_revenue * 100) if total_revenue else 0,
            "tax_rate": 25, "tax": tax, "pat": pat,
            "total_orders": len(orders), "total_sales": len(sales),
        }
    finally:
        db.close()


# ---- DASHBOARD ----
@app.get("/api/dashboard")
def dashboard():
    db = SessionLocal()
    try:
        all_sales = db.query(Sale).all()
        all_orders = db.query(Order).all()

        revenue = sum(s.total_amount or 0 for s in all_sales)
        freight = sum(s.freight_amount or 0 for s in all_sales)
        gp_total = sum(s.gp or 0 for s in all_sales)
        pending = sum(s.total_amount or 0 for s in all_sales if s.payment_status == "Pending")
        total_order_value = sum(o.invoice_amount or 0 for o in all_orders)
        total_order_cost = sum(o.value_excl_gst_freight or 0 for o in all_orders)

        recent_sales = []
        for s in db.query(Sale).order_by(Sale.id.desc()).limit(5).all():
            try:
                cust_name = ""
                if s.customer_id:
                    cust = db.query(Customer).filter(Customer.id == s.customer_id).first()
                    cust_name = cust.contact_name if cust else ""
                dt_str = ""
                if s.sale_date:
                    if hasattr(s.sale_date, 'strftime'):
                        dt_str = s.sale_date.strftime("%d %b %Y")
                    else:
                        dt_str = str(s.sale_date)[:10]
                recent_sales.append({
                    "id": s.id, "invoice": s.invoice_no or "",
                    "customer": s.party_name or cust_name or "",
                    "amount": s.total_amount or 0, "status": s.payment_status or "",
                    "date": dt_str
                })
            except Exception:
                continue

        recent_orders = []
        for o in db.query(Order).order_by(Order.id.desc()).limit(5).all():
            recent_orders.append({
                "id": o.id, "sl_no": o.sl_no or 0,
                "po_no": o.po_no or "", "billing_site": o.billing_site or "",
                "invoice_no": o.invoice_no or "",
                "invoice_amount": o.invoice_amount or 0,
                "entry_date": o.entry_date or ""
            })

        return {
            "total_products": db.query(Product).count(),
            "total_customers": db.query(Customer).count(),
            "total_orders": db.query(Order).count(),
            "total_sales": db.query(Sale).count(),
            "revenue": revenue,
            "freight": freight,
            "gp_total": gp_total,
            "pending": pending,
            "total_order_value": total_order_value,
            "total_order_cost": total_order_cost,
            "recent_sales": recent_sales,
            "recent_orders": recent_orders,
        }
    except Exception as e:
        return {"error": str(e), "total_products": 0, "total_customers": 0,
                "total_orders": 0, "total_sales": 0, "revenue": 0, "freight": 0,
                "gp_total": 0, "pending": 0, "total_order_value": 0,
                "total_order_cost": 0, "recent_sales": [], "recent_orders": []}
    finally:
        db.close()


# ---- SETTINGS ----
@app.get("/api/settings")
def get_settings():
    db = SessionLocal()
    try:
        rows = db.query(Settings).all()
        return {s.key: s.value for s in rows}
    finally:
        db.close()


@app.put("/api/settings")
def update_settings(body: dict):
    db = SessionLocal()
    try:
        for k, v in body.items():
            row = db.query(Settings).filter(Settings.key == k).first()
            if row:
                row.value = str(v)
            else:
                db.add(Settings(key=k, value=str(v)))
        db.commit()
        return {"message": "Settings updated"}
    finally:
        db.close()


# ---- CSV IMPORT ----
def parse_csv_amount(val):
    if not val or str(val).strip() in ('-', '–', '', 'None', 'none'):
        return 0
    return float(str(val).replace('₹', '').replace(',', '').replace(' ', '').strip() or 0)


def parse_csv_date(val):
    if not val or val.strip() in ('-', '–', ''):
        return None
    from dateutil import parser as dateparser
    try:
        return dateparser.parse(val.strip(), dayfirst=False).strftime('%Y-%m-%d')
    except Exception:
        return val.strip() if val else None


@app.post("/api/import/orders")
async def import_orders_csv(file: UploadFile = File(...)):
    content = await file.read()
    text = content.decode('utf-8-sig')
    import csv, io
    reader = csv.DictReader(io.StringIO(text))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            sl_raw = row.get('Sl No.', '').strip()
            if not sl_raw:
                skipped += 1
                continue
            try:
                sl_no = int(sl_raw)
            except ValueError:
                skipped += 1
                continue
            existing = db.query(Order).filter(Order.sl_no == sl_no).first()
            data = {
                "sl_no": sl_no,
                "po_no": "",
                "po_date": parse_csv_date(row.get('PO Date', '')),
                "customer_name": row.get('Customer Name', '').strip(),
                "billing_site": row.get('Billing Site', '').strip(),
                "shipping_site": row.get('Shipping Site', '').strip(),
                "no_of_boxes": int(parse_csv_amount(row.get('No. Of Boxes', '0'))),
                "value_excl_gst_freight": parse_csv_amount(row.get('Value (excl. GST & Freight)', '0')),
                "invoice_no": row.get('Invoice No.', '').strip().replace('-', '') if row.get('Invoice No.', '').strip() not in ('-', '–', '') else '',
                "invoice_date": parse_csv_date(row.get('Invoice Date', '')),
                "invoice_amount_excl_gst": parse_csv_amount(row.get('Invoice Amount (ex. GST)', '0')),
                "weight_kgs": parse_csv_amount(row.get('Weight (Kg)', '0')),
                "freight_rate_per_kg": parse_csv_amount(row.get('Freight (Rate / Kg)', '0')),
                "transport_charges": parse_csv_amount(row.get('Transport Charges', '0')),
                "invoice_amount": parse_csv_amount(row.get('Invoice Amount', '0')),
                "eway_bill_no": row.get('E-way Bill No', '').strip() if row.get('E-way Bill No', '').strip() not in ('-', '–', '') else '',
                "lr_no": row.get('LR Copy', '').strip() if row.get('LR Copy', '').strip() not in ('-', '–', '') else '',
                "entry_date": parse_csv_date(row.get('ERP Entry Date', '')),
                "credit_note_amount": parse_csv_amount(row.get('Credit Note Amount (If any)', '0')),
                "credit_note_no": row.get('Credit Note No.', '').strip() if row.get('Credit Note No.', '').strip() not in ('-', '–', '') else '',
                "transporter": row.get('Transporter', '').strip(),
                "transporter_no": "",
            }
            if existing:
                for k, v in data.items():
                    if k != "sl_no":
                        setattr(existing, k, v)
            else:
                o = Order(**data)
                db.add(o)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} orders, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@app.post("/api/import/sales")
async def import_sales_csv(file: UploadFile = File(...)):
    content = await file.read()
    text = content.decode('utf-8-sig')
    import csv, io
    reader = csv.DictReader(io.StringIO(text))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            sl_raw = row.get('Sl No.', '').strip()
            if not sl_raw:
                skipped += 1
                continue
            try:
                sl_no = int(sl_raw)
            except ValueError:
                skipped += 1
                continue
            invoice_no = row.get('Raksha Invoice NO', '').strip()
            if invoice_no in ('-', '–', ''):
                invoice_no = f"INDORE-{sl_no:04d}"
            sale_date = parse_csv_date(row.get('Date ', '') or row.get('Date', ''))
            sale_date_dt = None
            if sale_date:
                try:
                    from datetime import datetime as dt
                    sale_date_dt = dt.strptime(sale_date, '%Y-%m-%d')
                except Exception:
                    pass
            freight = parse_csv_amount(row.get('Freight', '0'))
            gp = parse_csv_amount(row.get('GP', '0'))
            gp_pct_raw = row.get('GP%', '0').replace('%', '').strip()
            try:
                gp_pct = float(gp_pct_raw) if gp_pct_raw and gp_pct_raw not in ('-', '–', '', 'None') else 0
            except ValueError:
                gp_pct = 0
            invoice_value = parse_csv_amount(row.get('Invoice Value', '') or row.get('invoice_value', '0'))
            s = Sale(
                invoice_no=invoice_no,
                sale_date=sale_date_dt,
                payment_terms=row.get('Payment Terms', '').strip(),
                party_name=row.get('Party Name ', '') or row.get('Party Name', '').strip(),
                location=row.get('Location', '').strip(),
                pincode=row.get('Pincode', '').strip(),
                state=row.get('State', '').strip(),
                transporter_name=row.get('Transporter Name', '').strip(),
                lr_no=row.get('LR No', '').strip(),
                freight_amount=freight,
                weight_kgs=parse_csv_amount(row.get('Weight', '0')),
                weight_pg_fiber=parse_csv_amount(row.get('Weight on PG Fiber Bill', '0')),
                sales_person=row.get('Sales Ex Person / Person In-Charge', '').strip(),
                pg_fiber_invoice_no=(row.get('P.G.Fiber Invoice No', '') or row.get('P.G.Fiber Invoice No ', '') or '').strip(),
                pg_fiber_invoice_value=parse_csv_amount(row.get('P.G.Fiber Invoice Value', '') or row.get('P.G.Fiber Invoice Value ', '0')),
                gp=gp,
                gp_percent=gp_pct,
                invoice_value=invoice_value,
                total_amount=freight,
                source_csv="From Indore",
            )
            db.add(s)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} sales, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


def map_csv_col(row, keys, default=""):
    for k in keys:
        v = row.get(k, "")
        if v and v.strip() and v.strip() not in ('-', '–'):
            return v.strip()
    return default


@app.post("/api/import/products")
async def import_products_csv(file: UploadFile = File(...)):
    content = await file.read()
    text = content.decode('utf-8-sig')
    import csv, io
    reader = csv.DictReader(io.StringIO(text))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            name = map_csv_col(row, ['Name', 'Description', 'Product Name', 'name', 'description'])
            if not name:
                skipped += 1
                continue
            part_no = map_csv_col(row, ['Part No', 'Part Number', 'SKU', 'part_no'])
            category = map_csv_col(row, ['Category', 'category'], 'Manhole Cover')
            size = map_csv_col(row, ['Size', 'size'])
            load_rating = map_csv_col(row, ['Load Rating', 'Load', 'load_rating'], '5 Ton')
            material = map_csv_col(row, ['Material', 'material'], 'FRP')
            color = map_csv_col(row, ['Color', 'color'], 'Grey')
            unit = map_csv_col(row, ['Unit', 'unit'], 'Nos')
            hsn_code = map_csv_col(row, ['HSN Code', 'HSN', 'hsn_code'])
            mrp = parse_csv_amount(map_csv_col(row, ['MRP', 'Price', 'mrp'], '0'))

            if part_no:
                existing = db.query(Product).filter(Product.part_no == part_no).first()
                if existing:
                    existing.name = name
                    existing.category = category
                    existing.size = size
                    existing.load_rating = load_rating
                    existing.material = material
                    existing.color = color
                    existing.unit = unit
                    existing.hsn_code = hsn_code
                    if mrp and existing.pricing:
                        existing.pricing.mrp = mrp
                        existing.pricing.raw_material_cost = mrp
                        existing.pricing.total_cost = mrp
                    imported += 1
                    continue

            p = Product(part_no=part_no, name=name, category=category,
                        size=size, load_rating=load_rating, material=material,
                        color=color, unit=unit, hsn_code=hsn_code)
            db.add(p)
            db.flush()
            db.add(Pricing(product_id=p.id, raw_material_cost=mrp, total_cost=mrp, mrp=mrp))
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} products, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@app.post("/api/import/customers")
async def import_customers_csv(file: UploadFile = File(...)):
    content = await file.read()
    text = content.decode('utf-8-sig')
    import csv, io
    reader = csv.DictReader(io.StringIO(text))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            customer_id = map_csv_col(row, ['Customer ID', 'ID', 'Cust ID', 'customer_id'])
            if not customer_id:
                skipped += 1
                continue
            existing = db.query(Customer).filter(Customer.customer_id == customer_id).first()
            data = {
                "customer_id": customer_id,
                "gstin": map_csv_col(row, ['GSTIN', 'GST Number', 'GST', 'gstin']),
                "billing_address": map_csv_col(row, ['Billing Address', 'Address', 'billing_address']),
                "shipping_address": map_csv_col(row, ['Shipping Address', 'shipping_address']),
                "state": map_csv_col(row, ['State', 'state']),
                "district": map_csv_col(row, ['District', 'district']),
                "city": map_csv_col(row, ['City', 'city']),
                "pincode": map_csv_col(row, ['Pincode', 'Pin Code', 'ZIP', 'pincode']),
                "contact_name": map_csv_col(row, ['Contact Name', 'Name', 'contact_name']),
                "contact_number": map_csv_col(row, ['Contact Number', 'Phone', 'Mobile', 'contact_number']),
                "contact_email": map_csv_col(row, ['Contact Email', 'Email', 'contact_email']),
                "exec_code": map_csv_col(row, ['Executive Code', 'Exec Code', 'exec_code']),
                "exec_name": map_csv_col(row, ['Executive Name', 'Exec Name', 'exec_name']),
                "exec_number": map_csv_col(row, ['Executive Number', 'Exec Number', 'exec_number']),
                "exec_email": map_csv_col(row, ['Executive Email', 'Exec Email', 'exec_email']),
            }
            if existing:
                for k, v in data.items():
                    if k != "customer_id":
                        setattr(existing, k, v)
            else:
                c = Customer(**data)
                db.add(c)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} customers, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@app.post("/api/import/transporters")
async def import_transporters_csv(file: UploadFile = File(...)):
    content = await file.read()
    text = content.decode('utf-8-sig')
    import csv, io
    reader = csv.DictReader(io.StringIO(text))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            transporter_id = map_csv_col(row, ['Transporter ID', 'ID', 'transporter_id'])
            name = map_csv_col(row, ['Name', 'Transporter Name', 'name'])
            if not transporter_id or not name:
                skipped += 1
                continue
            existing = db.query(Transporter).filter(Transporter.transporter_id == transporter_id).first()
            data = {
                "transporter_id": transporter_id,
                "name": name,
                "phone": map_csv_col(row, ['Phone', 'Mobile', 'phone']),
                "email": map_csv_col(row, ['Email', 'email']),
                "address": map_csv_col(row, ['Address', 'address']),
                "state": map_csv_col(row, ['State', 'state']),
                "district": map_csv_col(row, ['District', 'district']),
                "city": map_csv_col(row, ['City', 'city']),
                "pincode": map_csv_col(row, ['Pincode', 'Pin Code', 'pincode']),
                "gst_number": map_csv_col(row, ['GST Number', 'GSTIN', 'GST', 'gst_number']),
                "pan_number": map_csv_col(row, ['PAN Number', 'PAN', 'pan_number']),
                "contact_person": map_csv_col(row, ['Contact Person', 'contact_person']),
                "contact_number": map_csv_col(row, ['Contact Number', 'contact_number']),
            }
            if existing:
                for k, v in data.items():
                    if k != "transporter_id":
                        setattr(existing, k, v)
            else:
                t = Transporter(**data)
                db.add(t)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} transporters, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@app.post("/api/import/expenses")
async def import_expenses_csv(file: UploadFile = File(...)):
    content = await file.read()
    text = content.decode('utf-8-sig')
    import csv, io
    reader = csv.DictReader(io.StringIO(text))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            category = map_csv_col(row, ['Category', 'category'])
            amount_raw = map_csv_col(row, ['Amount', 'amount'], '0')
            amount = parse_csv_amount(amount_raw)
            if not category or amount == 0:
                skipped += 1
                continue
            date_str = map_csv_col(row, ['Date', 'Expense Date', 'expense_date'])
            dt = None
            if date_str:
                try:
                    from dateutil import parser as dateparser
                    dt = dateparser.parse(date_str, dayfirst=False)
                except Exception:
                    pass
            if not dt:
                dt = datetime.utcnow()
            e = Expense(
                category=category,
                description=map_csv_col(row, ['Description', 'description']),
                amount=amount,
                vendor=map_csv_col(row, ['Vendor', 'vendor']),
                expense_date=dt
            )
            db.add(e)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} expenses, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


# ---- FILE UPLOAD (Cloudinary) ----
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".pdf"}

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, "Only .jpg, .png, .pdf files allowed")
    try:
        content = await file.read()
        rtype = "raw" if ext == ".pdf" else "image"
        result = cloudinary.uploader.upload(
            content,
            folder="raksha_erp",
            resource_type=rtype
        )
        url = result["secure_url"]
        if rtype == "raw" and "/image/" in url:
            url = url.replace("/image/upload/", "/raw/upload/")
        return {"filename": result["public_id"], "url": url, "original": file.filename}
    except Exception as e:
        raise HTTPException(500, f"Upload failed: {str(e)}")


@app.post("/api/import-standard-packaging")
async def import_standard_packaging(file: UploadFile = File(...)):
    db = SessionLocal()
    try:
        content = await file.read()
        filename = (file.filename or "").lower()

        if filename.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
        else:
            text_content = content.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text_content))
            rows = list(reader)

        if not rows:
            raise HTTPException(400, "File is empty")

        header_row_idx = 0
        for i, row in enumerate(rows):
            row_lower = [str(c).strip().lower() for c in row]
            if any("part no" in c for c in row_lower):
                header_row_idx = i
                break

        header = [h.strip().lower() for h in rows[header_row_idx]]
        part_no_idx = None
        box_idx = None
        mrp_idx = None
        desc_idx = None
        size_idx = None
        grp_idx = None
        for i, h in enumerate(header):
            if "part no" in h:
                part_no_idx = i
            if h in ("box", "boxes") or "std pack" in h or "std pkg" in h:
                box_idx = i
            if h in ("gen", "mrp") or h == "gen":
                mrp_idx = i
            if "product spec" in h or "description" in h:
                desc_idx = i
            if "size" in h:
                size_idx = i
            if "item grp" in h or "item group" in h or "grp" in h:
                grp_idx = i
        if part_no_idx is None or box_idx is None:
            raise HTTPException(400, f"File must have 'Part No' and 'Box'/'Std Packing' columns. Found: {header}")

        desc_idx = desc_idx or 1
        size_idx = size_idx or 2
        grp_idx = grp_idx or 4

        created = 0
        updated = 0
        not_found = []
        for row in rows[header_row_idx+1:]:
            if len(row) <= max(part_no_idx, box_idx):
                continue
            part_no = row[part_no_idx].strip().replace(" ", "")
            box_val = row[box_idx].strip()
            if not part_no or not box_val:
                continue
            try:
                pieces = int(float(box_val))
            except ValueError:
                continue

            desc = str(row[desc_idx] or "").strip() if len(row) > desc_idx else ""
            size_raw = str(row[size_idx] or "").strip() if len(row) > size_idx else ""
            item_grp = str(row[grp_idx] or "").strip().upper() if len(row) > grp_idx else "FRP"

            size_mm = size_raw.lower().replace(" ", "").replace("x", "x")
            tonnage = "10 Ton" if "10 ton" in desc.lower() else "5 Ton"
            color = "White" if "white" in desc.lower() or part_no.upper().endswith("-WH") else "Grey"
            has_lock = "lock" in desc.lower() or part_no.upper().endswith("L") and "GRY" not in part_no.upper()
            has_hinges = "hinge" in desc.lower()
            category = "Gully Cover" if "gully" in desc.lower() or item_grp == "GULLY" else "Manhole Cover"

            product = db.query(Product).filter(Product.part_no == part_no).first()
            if product:
                product.pieces_per_box = pieces
                product.std_packaging = pieces
                if mrp_idx is not None and len(row) > mrp_idx:
                    mrp_val = row[mrp_idx].strip().replace(",", "").replace("₹", "").replace("?", "")
                    try:
                        mrp = float(mrp_val)
                        if mrp > 0:
                            pr = db.query(Pricing).filter(Pricing.product_id == product.id).first()
                            if pr:
                                pr.mrp = mrp
                            else:
                                db.add(Pricing(product_id=product.id, mrp=mrp, gst_rate=18))
                    except ValueError:
                        pass
                updated += 1
            else:
                if not desc:
                    continue
                new_p = Product(
                    part_no=part_no, name=desc, category=category,
                    size=size_mm, load_rating=tonnage, material="FRP",
                    color=color, hsn_code="39259090",
                    pieces_per_box=pieces, std_packaging=pieces
                )
                db.add(new_p)
                db.flush()
                mrp_val = 0
                if mrp_idx is not None and len(row) > mrp_idx:
                    try:
                        mrp_val = float(row[mrp_idx].strip().replace(",", "").replace("₹", "").replace("?", ""))
                    except ValueError:
                        pass
                db.add(Pricing(product_id=new_p.id, mrp=mrp_val, gst_rate=18))
                created += 1

        db.commit()
        return {"updated": updated, "created": created, "not_found": not_found, "total_rows": len(rows) - header_row_idx - 1}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


# ---- XLSX IMPORT (Orders & Sales) ----
def read_xlsx_sheet(file_content, sheet_name=None):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    return rows


def rows_to_csv_string(rows):
    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


@app.post("/api/import/orders-xlsx")
async def import_orders_xlsx(file: UploadFile = File(...)):
    content = await file.read()
    try:
        rows = read_xlsx_sheet(content)
        if not rows:
            raise HTTPException(400, "File is empty")
        csv_text = rows_to_csv_string(rows)
        text_content = csv_text
        reader = csv.DictReader(io.StringIO(text_content))
        db = SessionLocal()
        imported = 0
        skipped = 0
        try:
            for row in reader:
                sl_raw = row.get('Sl No.', '').strip()
                if not sl_raw:
                    skipped += 1
                    continue
                try:
                    sl_no = int(float(sl_raw))
                except ValueError:
                    skipped += 1
                    continue
                existing = db.query(Order).filter(Order.sl_no == sl_no).first()
                data = {
                    "sl_no": sl_no,
                    "po_no": "",
                    "po_date": parse_csv_date(row.get('PO Date', '')),
                    "customer_name": row.get('Customer Name', '').strip(),
                    "billing_site": row.get('Billing Site', '').strip(),
                    "shipping_site": row.get('Shipping Site', '').strip(),
                    "no_of_boxes": int(parse_csv_amount(row.get('No. Of Boxes', '0'))),
                    "value_excl_gst_freight": parse_csv_amount(row.get('Value (excl. GST & Freight)', '0')),
                    "invoice_no": row.get('Invoice No.', '').strip() if row.get('Invoice No.', '').strip() not in ('-', '–', '') else '',
                    "invoice_date": parse_csv_date(row.get('Invoice Date', '')),
                    "invoice_amount_excl_gst": parse_csv_amount(row.get('Invoice Amount (ex. GST)', '0')),
                    "weight_kgs": parse_csv_amount(row.get('Weight (Kg)', '0')),
                    "freight_rate_per_kg": parse_csv_amount(row.get('Freight (Rate / Kg)', '0')),
                    "transport_charges": parse_csv_amount(row.get('Transport Charges', '0')),
                    "invoice_amount": parse_csv_amount(row.get('Invoice Amount', '0')),
                    "eway_bill_no": row.get('E-way Bill No', '').strip() if row.get('E-way Bill No', '').strip() not in ('-', '–', '') else '',
                    "lr_no": row.get('LR Copy', '').strip() if row.get('LR Copy', '').strip() not in ('-', '–', '') else '',
                    "entry_date": parse_csv_date(row.get('ERP Entry Date', '')),
                    "credit_note_amount": parse_csv_amount(row.get('Credit Note Amount (If any)', '0')),
                    "credit_note_no": row.get('Credit Note No.', '').strip() if row.get('Credit Note No.', '').strip() not in ('-', '–', '') else '',
                    "transporter": row.get('Transporter', '').strip(),
                    "transporter_no": "",
                }
                if existing:
                    for k, v in data.items():
                        if k != "sl_no":
                            setattr(existing, k, v)
                else:
                    o = Order(**data)
                    db.add(o)
                imported += 1
            db.commit()
            return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} orders from XLSX"}
        except Exception as e:
            db.rollback()
            raise HTTPException(500, f"Import failed: {str(e)}")
        finally:
            db.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Failed to read XLSX: {str(e)}")


@app.post("/api/import/sales-xlsx")
async def import_sales_xlsx(file: UploadFile = File(...)):
    content = await file.read()
    try:
        rows = read_xlsx_sheet(content)
        if not rows:
            raise HTTPException(400, "File is empty")
        csv_text = rows_to_csv_string(rows)
        reader = csv.DictReader(io.StringIO(csv_text))
        db = SessionLocal()
        imported = 0
        skipped = 0
        try:
            for row in reader:
                sl_raw = row.get('Sl No.', '').strip()
                if not sl_raw:
                    skipped += 1
                    continue
                try:
                    int(float(sl_raw))
                except ValueError:
                    skipped += 1
                    continue
                invoice_no = row.get('Raksha Invoice NO', '').strip()
                if invoice_no in ('-', '–', ''):
                    invoice_no = None
                sale_date = parse_csv_date(row.get('Date ', '') or row.get('Date', ''))
                sale_date_dt = None
                if sale_date:
                    try:
                        sale_date_dt = datetime.strptime(sale_date, '%Y-%m-%d')
                    except Exception:
                        pass
                freight = parse_csv_amount(row.get('Freight', '0'))
                gp = parse_csv_amount(row.get('GP', '0'))
                gp_pct_raw = row.get('GP%', '0').replace('%', '').strip()
                try:
                    gp_pct = float(gp_pct_raw) if gp_pct_raw and gp_pct_raw not in ('-', '–', '', 'None') else 0
                except ValueError:
                    gp_pct = 0
                invoice_value = parse_csv_amount(row.get('Invoice Value', '') or row.get('invoice_value', '0'))
                s = Sale(
                    invoice_no=invoice_no,
                    sale_date=sale_date_dt,
                    payment_terms=row.get('Payment Terms', '').strip(),
                    party_name=(row.get('Party Name ', '') or row.get('Party Name', '') or '').strip(),
                    location=row.get('Location', '').strip(),
                    pincode=row.get('Pincode', '').strip(),
                    state=row.get('State', '').strip(),
                    transporter_name=row.get('Transporter Name', '').strip(),
                    lr_no=row.get('LR No', '').strip(),
                    freight_amount=freight,
                    weight_kgs=parse_csv_amount(row.get('Weight', '0')),
                    weight_pg_fiber=parse_csv_amount(row.get('Weight on PG Fiber Bill', '0')),
                    sales_person=row.get('Sales Ex Person / Person In-Charge', '').strip(),
                    pg_fiber_invoice_no=(row.get('P.G.Fiber Invoice No', '') or row.get('P.G.Fiber Invoice No ', '') or '').strip(),
                    pg_fiber_invoice_value=parse_csv_amount(row.get('P.G.Fiber Invoice Value', '') or row.get('P.G.Fiber Invoice Value ', '0')),
                    gp=gp,
                    gp_percent=gp_pct,
                    invoice_value=invoice_value,
                    total_amount=freight,
                    source_csv="From Indore",
                )
                db.add(s)
                imported += 1
            db.commit()
            return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} sales from XLSX"}
        except Exception as e:
            db.rollback()
            raise HTTPException(500, f"Import failed: {str(e)}")
        finally:
            db.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Failed to read XLSX: {str(e)}")


# ---- EXPORT (CSV, XLSX, PDF) ----
@app.get("/api/export/orders")
def export_orders(format: str = "csv"):
    db = SessionLocal()
    try:
        rows = db.query(Order).order_by(Order.sl_no).all()
        headers = ["Sl No.", "PO No.", "PO Date", "Customer Name", "Billing Site", "Shipping Site", "No. Of Boxes",
                    "Value (excl. GST & Freight)", "Invoice No.", "Invoice Date",
                    "Invoice Amount (ex. GST)", "Weight (Kg)", "Freight (Rate / Kg)",
                    "Transport Charges", "Invoice Amount", "E-way Bill No", "LR No.",
                    "Entry Date", "Credit Note Amount", "Credit Note No.", "Transporter"]
        data = []
        for o in rows:
            data.append([o.sl_no, o.po_no or "", o.po_date or "", o.customer_name or "", o.billing_site or "", o.shipping_site or "",
                         o.no_of_boxes or 0, o.value_excl_gst_freight or 0, o.invoice_no or "",
                         o.invoice_date or "", o.invoice_amount_excl_gst or 0, o.weight_kgs or 0,
                         o.freight_rate_per_kg or 0, o.transport_charges or 0, o.invoice_amount or 0,
                         o.eway_bill_no or "", o.lr_no or "", o.entry_date or "",
                         o.credit_note_amount or 0, o.credit_note_no or "", o.transporter or ""])

        if format == "xlsx":
            return export_xlsx("Orders", headers, data)
        elif format == "pdf":
            return export_pdf("Orders", headers, data)
        else:
            return export_csv(headers, data)
    finally:
        db.close()


@app.get("/api/export/sales")
def export_sales(format: str = "csv"):
    db = SessionLocal()
    try:
        rows = db.query(Sale).order_by(Sale.id.desc()).all()
        headers = ["Invoice No.", "Date", "Party Name", "Location", "State", "Transporter",
                    "Freight", "Weight", "Weight PG Fiber", "Invoice Value", "GP", "GP%",
                    "Payment Terms", "Sales Person", "PG Fiber Invoice No", "PG Fiber Invoice Value"]
        data = []
        for s in rows:
            dt = ""
            if s.sale_date:
                try:
                    dt = s.sale_date.strftime("%Y-%m-%d")
                except Exception:
                    dt = str(s.sale_date)[:10]
            data.append([s.invoice_no or "", dt, s.party_name or "", s.location or "",
                         s.state or "", s.transporter_name or "", s.freight_amount or 0,
                         s.weight_kgs or 0, s.weight_pg_fiber or 0, s.invoice_value or 0,
                         s.gp or 0, s.gp_percent or 0, s.payment_terms or "",
                         s.sales_person or "", s.pg_fiber_invoice_no or "",
                         s.pg_fiber_invoice_value or 0])

        if format == "xlsx":
            return export_xlsx("Sales", headers, data)
        elif format == "pdf":
            return export_pdf("Sales", headers, data)
        else:
            return export_csv(headers, data)
    finally:
        db.close()


@app.get("/api/export/proforma-orders")
def export_proforma_orders(format: str = "csv", order_type: str = None):
    db = SessionLocal()
    try:
        query = db.query(ProformaOrder)
        if order_type:
            query = query.filter(ProformaOrder.order_type == order_type)
        rows = query.order_by(ProformaOrder.created_at.desc()).all()
        headers = ["PI No", "Date", "Customer", "Type", "Billing Site", "Shipping Site",
                    "Boxes", "Total Qty", "Value (excl GST)", "GST", "Freight",
                    "Total Amount", "Payment Status", "Delivery Days"]
        data = []
        for o in rows:
            cust = db.query(Customer).filter(Customer.id == o.customer_id).first()
            data.append([o.pi_no or "", o.pi_date.strftime("%Y-%m-%d") if o.pi_date else "",
                         cust.contact_name if cust else "", o.order_type or "",
                         o.billing_site or "", o.shipping_site or "",
                         o.no_of_boxes or 0, o.total_qty or 0, o.value_excl_gst or 0,
                         o.gst_amount or 0, o.freight_amount or 0, o.total_amount or 0,
                         o.payment_status or "", o.delivery_days or 0])

        if format == "xlsx":
            return export_xlsx("PI-PO Orders", headers, data)
        elif format == "pdf":
            return export_pdf("PI-PO Orders", headers, data)
        else:
            return export_csv(headers, data)
    finally:
        db.close()


def export_csv(headers, data):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in data:
        writer.writerow(row)
    csv_bytes = output.getvalue().encode('utf-8-sig')
    return Response(content=csv_bytes, media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=export.csv"})


def export_xlsx(sheet_name, headers, data):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for row in data:
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={sheet_name}.xlsx"})


def export_pdf(title, headers, data):
    from fpdf import FPDF
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"Raksha ERP - {title}", ln=True, align="C")
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}", ln=True, align="C")
    pdf.ln(4)
    num_cols = len(headers)
    col_width = max(277 / num_cols, 20)
    pdf.set_font("Helvetica", "B", 7)
    for h in headers:
        short_h = str(h)[:20]
        pdf.cell(col_width, 7, short_h, border=1, align="C")
    pdf.ln()
    pdf.set_font("Helvetica", "", 6)
    for row in data:
        for val in row:
            s = str(val)[:22]
            pdf.cell(col_width, 5, s, border=1)
        pdf.ln()
    pdf_bytes = pdf.output()
    return Response(content=bytes(pdf_bytes), media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={title}.pdf"})


# ---- FRONTEND ----
frontend_path = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.isdir(frontend_path):
    app.mount("/static", StaticFiles(directory=frontend_path), name="static")


@app.get("/")
def index():
    index_path = os.path.join(frontend_path, "index.html")
    if os.path.isfile(index_path):
        return FileResponse(index_path)
    return {"message": "Raksha ERP API is running. Frontend not found."}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
